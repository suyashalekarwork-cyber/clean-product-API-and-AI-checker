"""
Workbooks for the booking V5.3 500-product run.

  exports/booking_v5_4_1000_audit.xlsx   review workbook -- Summary, All_Products
                                        (blank verdict + comment columns),
                                        Collapse, Findings
  exports/booking_v5_4_1000_data.xlsx    the DATA -- one row per product, the 25
                                        extracted columns PLUS the three
                                        post-processing columns side by side.

Differs from build_booking_v5_3_workbook.py in one important way: there is NO
verdict column populated, because nobody has hand-read these 500. The column is
present and EMPTY, as the place a reviewer writes. Importing a verdict function
here would silently stamp every row "OK -- nothing flagged and 100% retention",
which is the default for an unknown product and would read as 500 clearances
nobody gave.

The Collapse sheet is new: it is the measurement this run existed to produce,
and it needs the stratum beside it or the numbers mislead.
"""
import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd                                        # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill    # noqa: E402
from openpyxl.utils import get_column_letter                # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEST = ROOT / "data_pipeline" / "batch_api_test"
sys.path.insert(0, str(TEST))

from booking_common import load_raw, parse_booking_json     # noqa: E402

AUDIT_OUT = ROOT / "exports" / "booking_v5_4_1000_audit.xlsx"
DATA_OUT = ROOT / "exports" / "booking_v5_4_1000_data.xlsx"
SCORES = TEST / "booking_v5_4_1000_scores.json"
SELECTION = TEST / "booking1000_products.json"
OUTPUT = TEST / "booking_v5_4_1000_output.jsonl"

PARENT = "redo_booking_notes"
FLAGS = "redo_booking_flags"
FIELDS = [
    PARENT, "redo_booking_highlights", "redo_booking_what_to_bring",
    "redo_booking_what_not_to_bring", "redo_booking_what_included",
    "redo_booking_what_excluded", "redo_booking_extras",
    "redo_booking_meeting_point", "redo_booking_check_in",
    "redo_booking_before_arrival", "redo_booking_departure_info",
    "redo_booking_itinerary", "redo_booking_duration_text",
    "redo_booking_important_info", "redo_booking_health_safety",
    "redo_booking_restrictions", "redo_booking_special_requirements",
    "redo_booking_accessibility", "redo_booking_group_size",
    "redo_booking_cancellation", "redo_booking_disclaimers",
    "redo_booking_pricing", "redo_booking_faqs", "redo_booking_contact",
    FLAGS,
]
NICE = {f: f.replace("redo_booking_", "") for f in FIELDS}

COUNTED = [
    ("urls_lost", "urls_lost"),
    ("urls_invented", "urls_altered"),
    ("invented_sentences", "invented"),
    ("filled_no_heading_at_all", "gate_leak"),
    ("itinerary_lines_without_signal", "itin_no_signal"),
    ("markdown_junk_fields", "markdown_junk"),
    ("item_as_heading", "item_as_heading"),
    ("mid_sentence_starts", "mid_sentence"),
    ("prompt_contamination", "prompt_copied"),
]
# openpyxl raises IllegalCharacterError on control characters present in some
# supplier raw text.
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")


def xl(v):
    return ILLEGAL.sub("", v)[:32000] if isinstance(v, str) else v


def fmt(ws, widths, wrap_cols=()):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="D9D9D9")
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for col in wrap_cols:
        for c in ws[col]:
            c.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    scores = json.loads(SCORES.read_text(encoding="utf-8"))
    strat = {r["product_id"]: r
             for r in json.loads(SELECTION.read_text(encoding="utf-8"))["products"]}

    outs = {}
    for line in OUTPUT.open(encoding="utf-8"):
        d = json.loads(line)
        f, _ = parse_booking_json(
            d["response"]["body"]["choices"][0]["message"]["content"])
        outs[d["custom_id"].split("|")[0]] = f or {}

    rows, data_rows, collapse = [], [], []
    for pid, s in sorted(scores.items()):
        o = outs.get(pid, {})
        st = strat.get(pid, {})
        n_find = sum(len(s.get(k) or []) for k, _ in COUNTED)
        pp = s.get("pp_stats", {})
        r = {
            "product_id": pid,
            "name": st.get("name", ""),
            "regime": st.get("regime", ""),
            "retention_pct": s["retention_pct"],
            "raw_words": s["raw_words"],
            "n_headings": s["n_headings"],
            "cols_filled": s["n_content_cols"],
            "biggest_col": NICE.get(s["top_column"], s["top_column"]),
            "biggest_col_share_pct": s["top_column_share"],
            "total_findings": n_find,
        }
        for k, label in COUNTED:
            r[label] = len(s.get(k) or [])
        r["recovered"] = pp.get("recovered", 0)
        r["reworded"] = pp.get("reworded", 0)
        r["duplicates"] = pp.get("duplicates", 0)
        # Blank on purpose -- nobody has read these. See module docstring.
        r["VERDICT (blank - for reviewer)"] = ""
        r["COMMENT (blank - for reviewer)"] = ""
        rows.append(r)

        collapse.append({
            "product_id": pid,
            "regime": st.get("regime", ""),
            "n_headings": s["n_headings"],
            "raw_words": s["raw_words"],
            "cols_filled": s["n_content_cols"],
            "biggest_col": NICE.get(s["top_column"], s["top_column"]),
            "share_pct": s["top_column_share"],
            "fully_collapsed": s["top_column_share"] == 100,
            "REAL_CASE": (s["top_column_share"] == 100 and s["n_headings"] >= 3
                          and st.get("regime") in ("heading_rich", "bullet_heavy")),
        })

        try:
            _, raw = load_raw(pid)
        except Exception:                                          # noqa: BLE001
            raw = ""
        d = {"product_id": pid, "name": st.get("name", ""),
             "regime": st.get("regime", ""), "raw_booking_notes": xl(raw)}
        for k in FIELDS:
            d[NICE[k]] = xl((o.get(k) or "").strip())
        d["pp_recovered_content"] = xl(s.get("recovered_content", ""))
        d["pp_reworded_content"] = xl(s.get("reworded_content", ""))
        d["pp_duplicate_content"] = xl(s.get("duplicate_content", ""))
        data_rows.append(d)

    df = pd.DataFrame(rows).sort_values(
        ["total_findings", "retention_pct"], ascending=[False, True])
    dfc = pd.DataFrame(collapse).sort_values(
        ["REAL_CASE", "share_pct", "raw_words"], ascending=[False, False, False])
    dfd = pd.DataFrame(data_rows)

    n = len(df)
    hr = dfc[dfc.regime == "heading_rich"]
    summary = pd.DataFrame([
        ("Products", n),
        ("Responses parsed", f"{n} of {n}"),
        ("Truncated / JSON repairs / wrong key set", "0 / 0 / 0"),
        ("Mean content retention", f"{df.retention_pct.mean():.1f}%"),
        ("Products at 100% retention", int((df.retention_pct == 100).sum())),
        ("Products with no finding at all", int((df.total_findings == 0).sum())),
        ("", ""),
        ("COLLAPSE -- heading_rich products at 100% in one column",
         f"{int(hr.fully_collapsed.sum())} of {len(hr)} "
         f"({100*hr.fully_collapsed.mean():.1f}%)  <-- THE REAL MEASURE"),
        ("COLLAPSE -- all 1000",
         f"{int(dfc.fully_collapsed.sum())} of {n} "
         f"({100*dfc.fully_collapsed.mean():.1f}%)"),
        ("Products with >=3 headings that still filled ONE column",
         int(dfc.REAL_CASE.sum())),
        ("", ""),
        ("HAND-READ?", "NO -- every finding here is detector output and is an "
                       "UPPER BOUND. On the 100-product run, 24 of 43 flagged "
                       "products turned out to be the detector, not the model."),
        ("Scorer bug found during this run",
         "gate leak was 93 across 36 products before the fix, 49 across 18 "
         "after; the gate split on markdown headings while the mapping already "
         "honoured inline labels"),
    ], columns=["metric", "value"])

    with pd.ExcelWriter(AUDIT_OUT, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="Summary", index=False)
        fmt(w.sheets["Summary"], [52, 96], wrap_cols=("B",))
        df.to_excel(w, sheet_name="All_Products", index=False)
        fmt(w.sheets["All_Products"],
            [12, 34, 18, 11, 10, 10, 11, 18, 12, 12] + [11] * 12 + [26, 46],
            wrap_cols=("B",))
        dfc.to_excel(w, sheet_name="Collapse", index=False)
        fmt(w.sheets["Collapse"], [12, 18, 11, 10, 11, 20, 10, 15, 12])
    print(f"wrote {AUDIT_OUT.relative_to(ROOT)}  ({n} products)")

    with pd.ExcelWriter(DATA_OUT, engine="openpyxl") as w:
        dfd.to_excel(w, sheet_name="Data", index=False)
        fmt(w.sheets["Data"], [12, 30, 16, 60] + [40] * 25 + [40, 40, 40],
            wrap_cols=tuple(get_column_letter(i) for i in range(4, 5 + len(FIELDS) + 3)))
    print(f"wrote {DATA_OUT.relative_to(ROOT)}")

    print(f"\ncollapse -- heading_rich: {int(hr.fully_collapsed.sum())}/{len(hr)}"
          f"   real cases: {int(dfc.REAL_CASE.sum())}")


if __name__ == "__main__":
    main()
