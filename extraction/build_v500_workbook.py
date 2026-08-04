"""
Isolated POC: Step 8 -- build v500_products.xlsx. Same format as
v50_products.xlsx, but Content_PreFix is skipped at this scale (pre-fix
state stays in v500_pre_fix_state.json), and a new Failures sheet is added
for products dropped in Step 2 (batch failures).

Usage:
    python build_v500_workbook.py
"""
import sys
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

TEST_DIR = Path(__file__).resolve().parent
OUT_PATH = TEST_DIR / "v500_products.xlsx"

DESC_FIELDS = [
    "redo_desc_about", "redo_desc_highlights", "redo_desc_what_included",
    "redo_desc_what_excluded", "redo_desc_itinerary", "redo_desc_what_to_bring",
    "redo_desc_duration_text", "redo_desc_requirements",
    "redo_desc_cancellation", "redo_desc_check_in",
    "redo_min_age", "redo_max_age", "redo_group_size", "redo_meeting_point",
    "redo_desc_other",
]
BOOKING_FIELDS = [
    "redo_booking_what_to_bring", "redo_booking_what_not_to_bring", "redo_booking_inclusions",
    "redo_booking_location", "redo_booking_check_in", "redo_booking_departure_info",
    "redo_booking_itinerary", "redo_booking_important_info", "redo_booking_cancellation",
    "redo_booking_faqs", "redo_booking_before_arrival", "redo_booking_contact",
    "redo_booking_other",
]
ALL_FIELDS = DESC_FIELDS + BOOKING_FIELDS

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_sheet(worksheet, n_cols, col_widths=None, wrap_cols=None):
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    if col_widths:
        for col_idx, width in col_widths.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment = Alignment(vertical="top")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if wrap_cols and cell.column in wrap_cols:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = top_alignment


def main():
    pre_fix = json.loads((TEST_DIR / "v500_pre_fix_state.json").read_text(encoding="utf-8"))
    post_fix = json.loads((TEST_DIR / "v500_post_fix_state.json").read_text(encoding="utf-8"))
    selection_df = pd.read_csv(TEST_DIR / "new_500_products_selection.csv", dtype={"product_id": str}).set_index("product_id")
    rescreen_df = pd.read_csv(TEST_DIR / "rescreen_v500_results.csv", dtype={"product_id": str}).set_index("product_id")
    paste_df = pd.read_csv(TEST_DIR / "code_only_fix_v500_paste_records.csv", dtype={"product_id": str})
    partial_df = pd.read_csv(TEST_DIR / "code_only_fix_v500_partial_records.csv", dtype={"product_id": str})
    assertions_df = pd.read_csv(TEST_DIR / "assertions_v500_results.csv")

    failures_path = TEST_DIR / "v500_failures.csv"
    dropped_path = TEST_DIR / "v500_screen_dropped.csv"
    def safe_read_csv(path, **kwargs):
        if not path.exists():
            return None
        try:
            return pd.read_csv(path, **kwargs)
        except pd.errors.EmptyDataError:
            return None

    failures_df = safe_read_csv(failures_path)
    if failures_df is None:
        failures_df = pd.DataFrame(columns=["custom_id", "reason"])
    dropped_df = safe_read_csv(dropped_path, dtype={"product_id": str})
    if dropped_df is None:
        dropped_df = pd.DataFrame(columns=["product_id", "reason"])

    product_ids = list(pre_fix.keys())

    # ---- Sheet 1: Summary ----
    summary_rows = []
    for pid in product_ids:
        sel_row = selection_df.loc[pid] if pid in selection_df.index else None
        rs_row = rescreen_df.loc[pid] if pid in rescreen_df.index else None
        n_pasted = (paste_df[(paste_df["product_id"] == pid) & (paste_df["action"] == "pasted")]).shape[0]

        missing_after = rs_row["missing_after"] if rs_row is not None else 0
        status = "complete" if missing_after == 0 else "needs_review"
        flag_review = n_pasted >= 4

        summary_rows.append({
            "product_id": pid,
            "supplier_alias": sel_row["supplier_alias"] if sel_row is not None else "",
            "input_words": sel_row["input_words"] if sel_row is not None else 0,
            "band": sel_row["band"] if sel_row is not None else "",
            "missing_before": rs_row["missing_before"] if rs_row is not None else 0,
            "missing_after": missing_after,
            "coverage_before": rs_row["coverage_before"] if rs_row is not None else 100.0,
            "coverage_after": rs_row["coverage_after"] if rs_row is not None else 100.0,
            "sentences_pasted": n_pasted,
            "flag_for_human_review": flag_review,
            "extraction_status": status,
        })
    summary_df = pd.DataFrame(summary_rows)
    summary_df["product_id_sort"] = summary_df["product_id"].astype(int)
    summary_df = summary_df.sort_values("product_id_sort").drop(columns=["product_id_sort"]).reset_index(drop=True)

    # ---- Sheet 2: Paste_Detail ----
    paste_detail_df = paste_df[paste_df["action"].isin(["pasted", "skipped_duplicate"])].copy()
    paste_detail_df = paste_detail_df[["product_id", "sentence", "source_side", "home_bucket"]]
    if not paste_detail_df.empty:
        paste_detail_df["product_id_sort"] = paste_detail_df["product_id"].astype(int)
        paste_detail_df = paste_detail_df.sort_values("product_id_sort").drop(columns=["product_id_sort"]).reset_index(drop=True)

    # ---- Sheet 3: Partial_Recorded ----
    partial_sheet_df = partial_df[["product_id", "unit_text", "coverage_pct", "missing_phrase"]].copy()
    if not partial_sheet_df.empty:
        partial_sheet_df["product_id_sort"] = partial_sheet_df["product_id"].astype(int)
        partial_sheet_df = partial_sheet_df.sort_values("product_id_sort").drop(columns=["product_id_sort"]).reset_index(drop=True)

    # ---- Sheet 4: Content_PostFix ----
    content_post_rows = []
    for pid in product_ids:
        row = {"product_id": pid}
        row.update({f: post_fix[pid]["field_values"].get(f, "") for f in ALL_FIELDS})
        content_post_rows.append(row)
    content_post_df = pd.DataFrame(content_post_rows)
    content_post_df["product_id_sort"] = content_post_df["product_id"].astype(int)
    content_post_df = content_post_df.sort_values("product_id_sort").drop(columns=["product_id_sort"]).reset_index(drop=True)

    # ---- Sheet 5: Assertions ----
    assertions_sheet_df = assertions_df.copy()

    # ---- Sheet 6: Review_Queue (only products with 4+ pasted) ----
    flagged_pids = summary_df[summary_df["flag_for_human_review"]]["product_id"].tolist()
    review_rows = []
    for pid in flagged_pids:
        n_pasted = summary_df.loc[summary_df["product_id"] == pid, "sentences_pasted"].iloc[0]
        raw_desc = pre_fix[pid]["raw_desc"]
        raw_booking = pre_fix[pid]["raw_booking"]
        pasted_sentences = paste_df[(paste_df["product_id"] == pid) & (paste_df["action"] == "pasted")]
        pasted_summary = "\n".join(f"[{row['home_bucket']}] {row['sentence']}" for _, row in pasted_sentences.iterrows())

        review_rows.append({
            "product_id": pid,
            "sentences_pasted": n_pasted,
            "raw_description": raw_desc,
            "raw_booking_notes": raw_booking,
            "pasted_sentences_with_bucket": pasted_summary,
        })
    review_queue_df = pd.DataFrame(review_rows)
    if review_queue_df.empty:
        review_queue_df = pd.DataFrame(columns=["product_id", "sentences_pasted", "raw_description", "raw_booking_notes", "pasted_sentences_with_bucket"])
    else:
        review_queue_df["product_id_sort"] = review_queue_df["product_id"].astype(int)
        review_queue_df = review_queue_df.sort_values("product_id_sort").drop(columns=["product_id_sort"]).reset_index(drop=True)

    # ---- Sheet 7: Failures ----
    failures_sheet_rows = []
    for _, row in failures_df.iterrows():
        failures_sheet_rows.append({"identifier": row.get("custom_id", ""), "reason": row.get("reason", ""), "stage": "batch_extraction"})
    for _, row in dropped_df.iterrows():
        failures_sheet_rows.append({"identifier": row.get("product_id", ""), "reason": row.get("reason", ""), "stage": "screen_drop"})
    failures_sheet_df = pd.DataFrame(failures_sheet_rows) if failures_sheet_rows else pd.DataFrame(columns=["identifier", "reason", "stage"])

    # ---- write workbook ----
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        paste_detail_df.to_excel(writer, index=False, sheet_name="Paste_Detail")
        partial_sheet_df.to_excel(writer, index=False, sheet_name="Partial_Recorded")
        content_post_df.to_excel(writer, index=False, sheet_name="Content_PostFix")
        assertions_sheet_df.to_excel(writer, index=False, sheet_name="Assertions")
        review_queue_df.to_excel(writer, index=False, sheet_name="Review_Queue")
        failures_sheet_df.to_excel(writer, index=False, sheet_name="Failures")

        ws1 = writer.sheets["Summary"]
        style_sheet(ws1, len(summary_df.columns),
                    col_widths={1: 12, 2: 24, 3: 12, 4: 12, 5: 14, 6: 14, 7: 14, 8: 14, 9: 16, 10: 18, 11: 16},
                    wrap_cols={2})
        status_idx = summary_df.columns.get_loc("extraction_status") + 1
        flag_idx = summary_df.columns.get_loc("flag_for_human_review") + 1
        for r_idx, row in enumerate(summary_df.itertuples(), start=2):
            status = getattr(row, "extraction_status")
            flagged = getattr(row, "flag_for_human_review")
            status_cell = ws1.cell(row=r_idx, column=status_idx)
            status_cell.fill = GREEN_FILL if status == "complete" else RED_FILL
            if flagged:
                ws1.cell(row=r_idx, column=flag_idx).fill = AMBER_FILL

        ws2 = writer.sheets["Paste_Detail"]
        style_sheet(ws2, len(paste_detail_df.columns), col_widths={1: 12, 2: 70, 3: 16, 4: 22}, wrap_cols={2})

        ws3 = writer.sheets["Partial_Recorded"]
        if not partial_sheet_df.empty:
            style_sheet(ws3, len(partial_sheet_df.columns), col_widths={1: 12, 2: 60, 3: 14, 4: 50}, wrap_cols={2, 4})
        else:
            style_sheet(ws3, 4, col_widths={1: 12, 2: 60, 3: 14, 4: 50})

        ws4 = writer.sheets["Content_PostFix"]
        wrap_cols_4 = set(range(2, len(content_post_df.columns) + 1))
        style_sheet(ws4, len(content_post_df.columns), col_widths={i: (12 if i == 1 else 55) for i in range(1, len(content_post_df.columns) + 1)}, wrap_cols=wrap_cols_4)

        ws5 = writer.sheets["Assertions"]
        style_sheet(ws5, len(assertions_sheet_df.columns), col_widths={1: 25, 2: 16, 3: 12})
        status_idx5 = assertions_sheet_df.columns.get_loc("status") + 1
        for r_idx, row in enumerate(assertions_sheet_df.itertuples(), start=2):
            status = getattr(row, "status")
            cell = ws5.cell(row=r_idx, column=status_idx5)
            cell.fill = GREEN_FILL if status == "PASS" else RED_FILL

        ws6 = writer.sheets["Review_Queue"]
        style_sheet(ws6, len(review_queue_df.columns), col_widths={1: 12, 2: 16, 3: 60, 4: 60, 5: 70}, wrap_cols={3, 4, 5})

        ws7 = writer.sheets["Failures"]
        style_sheet(ws7, len(failures_sheet_df.columns), col_widths={1: 30, 2: 60, 3: 16}, wrap_cols={2})

    print(f"Wrote workbook to {OUT_PATH}")
    print(f"Sheets: Summary ({len(summary_df)}), Paste_Detail ({len(paste_detail_df)}), "
          f"Partial_Recorded ({len(partial_sheet_df)}), Content_PostFix ({len(content_post_df)}), "
          f"Assertions ({len(assertions_sheet_df)}), Review_Queue ({len(review_queue_df)}), "
          f"Failures ({len(failures_sheet_df)})")
    print(f"Products flagged for human review (4+ pasted): {len(flagged_pids)}")

    return summary_df, review_queue_df, failures_sheet_df


if __name__ == "__main__":
    main()
