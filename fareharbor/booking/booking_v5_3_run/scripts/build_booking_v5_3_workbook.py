"""
Two workbooks for the booking V5.3 run.

  exports/booking_v5_3_audit.xlsx   review workbook -- Summary, Findings,
                                    All_Products (verdict + comment), Per_Product
                                    (raw beside all 25 columns)
  exports/booking_v5_3_data.xlsx    the DATA -- one row per product, 25 extracted
                                    columns PLUS the three post-processing
                                    columns side by side. This is the closest
                                    thing to what a downstream consumer receives.

The post-processing columns are the point of the second file: recovered_content
and duplicate_content live in the scores JSON and the audit txt, but nowhere in
a machine-usable per-product table. Without that, "here is what we extracted"
and "here is what we missed and where it belonged" cannot be read together.
"""
import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd                                        # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill    # noqa: E402
from openpyxl.utils import get_column_letter                # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEST = ROOT / "data_pipeline" / "batch_api_test"
sys.path.insert(0, str(TEST))

from booking_common import load_raw, parse_booking_json     # noqa: E402
from audit_booking_v5_3_comments import verdict             # noqa: E402

AUDIT_OUT = ROOT / "exports" / "booking_v5_3_audit.xlsx"
DATA_OUT = ROOT / "exports" / "booking_v5_3_data.xlsx"

PARENT = "redo_booking_notes"
FLAGS = "redo_booking_flags"
FIELDS = [
    PARENT, "redo_booking_highlights", "redo_booking_what_to_bring",
    "redo_booking_what_not_to_bring", "redo_booking_what_included",
    "redo_booking_what_excluded", "redo_booking_extras",
    "redo_booking_meeting_point", "redo_booking_check_in",
    "redo_booking_before_arrival", "redo_booking_departure_info",
    "redo_booking_itinerary", "redo_booking_duration_text",
    "redo_booking_important_info", "redo_booking_health_safety",
    "redo_booking_restrictions", "redo_booking_special_requirements",
    "redo_booking_accessibility", "redo_booking_group_size",
    "redo_booking_cancellation", "redo_booking_disclaimers",
    "redo_booking_pricing", "redo_booking_faqs", "redo_booking_contact",
    FLAGS,
]
NICE = {f: f.replace("redo_booking_", "") for f in FIELDS}

FINDINGS = [
    ("prompt_contamination", "contamination"), ("urls_lost", "url_lost"),
    ("urls_invented", "url_altered"), ("invented_sentences", "invention"),
    ("mid_sentence_starts", "mid_sentence"), ("item_as_heading", "item_as_heading"),
    ("itinerary_lines_without_signal", "itinerary_line_test"),
    ("included_that_are_purchasable", "included_line_test"),
    ("markdown_junk_fields", "markdown_junk"), ("separators_kept", "separator_kept"),
]

# openpyxl refuses ASCII control characters that appear in some supplier text.
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")
CELL_LIMIT = 32000


def xl(v):
    if not isinstance(v, str):
        return v
    return ILLEGAL.sub("", v)[:CELL_LIMIT]


def load_outputs():
    out = {}
    for line in (TEST / "booking_v5_3_100_output.jsonl").open(encoding="utf-8"):
        d = json.loads(line)
        f, _ = parse_booking_json(
            d["response"]["body"]["choices"][0]["message"]["content"])
        out[d["custom_id"].split("|")[0]] = f or {}
    return out


def style(ws, widths, wrap_from=0):
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="D9D9D9")
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for i, c in enumerate(row):
            if i >= wrap_from:
                c.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    scores = json.loads((TEST / "booking_v5_3_100_scores.json").read_text(encoding="utf-8"))
    strat = {r["product_id"]: r for r in json.loads(
        (TEST / "booking100_products.json").read_text(encoding="utf-8"))["products"]}
    outs = load_outputs()

    data_rows, audit_rows, per_rows, finding_rows = [], [], [], []

    for pid, s in sorted(scores.items()):
        name, raw = load_raw(pid)
        o = outs.get(pid, {})
        code, comment = verdict(pid)

        # ---- DATA sheet: the 25 columns + the 3 post-processing columns ----
        row = {"product_id": pid, "product_name": name}
        for f in FIELDS:
            row[NICE[f]] = xl((o.get(f) or "").strip())
        row["recovered_content"] = xl(s.get("recovered_content", ""))
        row["reworded_content"] = xl(s.get("reworded_content", ""))
        row["duplicate_content"] = xl(s.get("duplicate_content", ""))
        data_rows.append(row)

        # ---- audit summary ----
        found = [lbl for k, lbl in FINDINGS if s.get(k)]
        audit_rows.append({
            "product_id": pid, "product_name": xl(name),
            "verdict": code, "comment": xl(comment),
            "stratum": strat.get(pid, {}).get("stratum", ""),
            "raw_words": s["raw_words"], "headings": s["n_headings"],
            "retention_pct": s["retention_pct"],
            "columns_filled": s["n_filled"],
            "findings": ", ".join(found),
            "recovered_n": s["pp_stats"]["recovered"],
            "reworded_n": s["pp_stats"]["reworded"],
            "duplicate_n": s["pp_stats"]["duplicates"],
        })

        for k, lbl in FINDINGS:
            for item in (s.get(k) or []):
                finding_rows.append({"product_id": pid, "finding": lbl,
                                     "detail": xl(str(item))})

        per_rows.append({"product_id": pid, "product_name": xl(name),
                         "raw_booking_notes": xl(raw),
                         **{NICE[f]: xl((o.get(f) or "").strip()) for f in FIELDS}})

    df_data = pd.DataFrame(data_rows)
    df_audit = pd.DataFrame(audit_rows)
    df_find = pd.DataFrame(finding_rows)
    df_per = pd.DataFrame(per_rows)

    # ---------- DATA workbook ----------
    with pd.ExcelWriter(DATA_OUT, engine="openpyxl") as w:
        df_data.to_excel(w, sheet_name="Data", index=False)
        style(w.sheets["Data"], [12, 34] + [40] * (len(df_data.columns) - 2), 2)
    print(f"wrote {DATA_OUT.relative_to(ROOT)}  "
          f"({len(df_data)} products x {len(df_data.columns)} columns)")

    # ---------- AUDIT workbook ----------
    counts = df_audit.verdict.value_counts()
    summary = pd.DataFrame({
        "metric": [
            "products", "mean retention %", "products at 100% retention",
            "--- verdicts ---", *[f"verdict: {k}" for k in counts.index],
            "--- gates ---",
            "text copied from the prompt", "URLs lost", "URLs altered",
            "invented sentences", "gate leaks (no heading at all)",
            "--- post-processing (report only) ---",
            "recovered_content entries", "reworded entries",
            "duplicate entries",
        ],
        "value": [
            len(df_audit), round(df_audit.retention_pct.mean(), 1),
            int((df_audit.retention_pct == 100).sum()),
            "", *[int(v) for v in counts.values],
            "",
            sum(len(s.get("prompt_contamination") or []) for s in scores.values()),
            sum(len(s.get("urls_lost") or []) for s in scores.values()),
            sum(len(s.get("urls_invented") or []) for s in scores.values()),
            sum(len(s.get("invented_sentences") or []) for s in scores.values()),
            sum(len(s.get("filled_no_heading_at_all") or []) for s in scores.values()),
            "",
            int(df_audit.recovered_n.sum()), int(df_audit.reworded_n.sum()),
            int(df_audit.duplicate_n.sum()),
        ],
    })

    with pd.ExcelWriter(AUDIT_OUT, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="Summary", index=False)
        style(w.sheets["Summary"], [44, 22])
        df_find.to_excel(w, sheet_name="Findings", index=False)
        style(w.sheets["Findings"], [12, 22, 110], 2)
        df_audit.to_excel(w, sheet_name="All_Products", index=False)
        style(w.sheets["All_Products"], [12, 34, 15, 90, 16, 10, 9, 12, 12, 30, 11, 11, 11], 3)
        df_per.to_excel(w, sheet_name="Per_Product", index=False)
        style(w.sheets["Per_Product"], [12, 30, 70] + [38] * len(FIELDS), 2)
    print(f"wrote {AUDIT_OUT.relative_to(ROOT)}  "
          f"(Summary / Findings {len(df_find)} / All_Products {len(df_audit)} / "
          f"Per_Product {len(df_per)})")


if __name__ == "__main__":
    main()
