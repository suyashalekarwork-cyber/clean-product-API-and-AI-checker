"""
Review workbook for the V5.3 500-product run.

Sheets:
  Priority_Matrix  P1-P4 with product IDs, impact and fix -- the same matrix as
                   the folder README, so the two cannot drift apart
  Issues_Only      the 25 flagged products, worst class first
  All_Products     all 499 with verdict + comment
  Per_Product      raw description beside all 22 extracted columns

Written to the repo folder, not exports/, since it ships with the run.
"""
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
TEST_DIR = ROOT / "data_pipeline" / "batch_api_test"
REPO = Path.home() / "Desktop" / "clean-product-API-and-AI-checker" / "v5_3_500_run"
sys.path.insert(0, str(TEST_DIR))

from build_model_comparison_batches import strip_html, find_raw_file  # noqa: E402
from audit_v5_3_500_comments import verdict as audit_verdict  # noqa: E402

OUT = REPO / "v5_3_500_audit.xlsx"
FIELDS = [
    "redo_desc_about", "redo_desc_important_info", "redo_desc_highlights",
    "redo_desc_what_included", "redo_desc_what_excluded", "redo_desc_extras",
    "redo_desc_itinerary", "redo_desc_what_to_bring", "redo_desc_duration_text",
    "redo_desc_cancellation", "redo_desc_check_in", "redo_desc_accessibility",
    "redo_desc_restrictions", "redo_desc_special_requirements",
    "redo_desc_faqs", "redo_desc_pricing", "redo_desc_disclaimers",
    "redo_meeting_point", "redo_min_age", "redo_max_age", "redo_group_size",
]
NICE = {f: f.replace("redo_desc_", "").replace("redo_", "") for f in FIELDS}

# One source of truth for the matrix -- the README quotes these same rows.
MATRIX = [
    ("CONTENT LOSS", "HIGH", 1, "535701",
     "\"Please order your protein at the time of booking\" -- an instruction the customer must act on",
     "Post-extraction check that returns orphaned sentences to about", "ours"),
    ("CONTENT LOSS", "MEDIUM", 1, "371805",
     "Engine spec dropped -- happens every run, the only repeatable loss",
     "Post-extraction check", "ours"),
    ("CONTENT LOSS", "LOW", 1, "293135", "Opening tagline dropped",
     "Post-extraction check", "ours"),

    ("MISCLASSIFICATION", "HIGH", 1, "634003",
     "'Departure times' and 'Arrival times' both dropped, the two lists merged -- Mission Beach "
     "5:30 PM reads as a fourth pickup point",
     "Contrasting label pairs must survive even where single lead-ins are dropped", "ours"),
    ("MISCLASSIFICATION", "MEDIUM", 2, "466438, 491113",
     "Difficulty rating in the Restrictions box -- reads 'Moderate' / 'Level: Hard'",
     "Two lines of prompt: difficulty names no column -> about", "ours"),
    ("MISCLASSIFICATION", "MEDIUM", 1, "639882",
     "Sub-labels 'Scenic Landscapes:', 'Wildlife Encounters:', 'Dress Code:' dropped",
     "Same label-pair rule as 634003", "ours"),
    ("MISCLASSIFICATION", "LOW", 3, "713497, 324361, 697755",
     "Label stripped -- every value survived, the section just has no title",
     "Tighten the inline-label rule", "ours"),
    ("MISCLASSIFICATION", "LOW", 1, "251713",
     "Content overrode the heading (the answer is arguably better)",
     "Watch only -- do not fix yet", "ours"),
    ("MISCLASSIFICATION", "LOW", 2, "198064, 501920",
     "A marketing line sitting in a list column", "Low value -- leave", "ours"),

    ("SUPPLIER MISTAKE", "HIGH", 3, "564767, 531290, 598043",
     "No description at all -- the raw is just the product name",
     "Chase the supplier; no prompt can fix this", "supplier"),
    ("SUPPLIER MISTAKE", "HIGH", 1, "327258",
     "What-to-Bring says the supplier PROVIDES everything -- the opposite",
     "Phase 2: extend the line test to what_to_bring", "supplier"),
    ("SUPPLIER MISTAKE", "HIGH", 1, "266189",
     "The whole description is the string 'meeting_point: Te Anau'",
     "Nothing to do -- reproduced faithfully", "supplier"),
    ("SUPPLIER MISTAKE", "MEDIUM", 2, "156525, 500245",
     "Notes / cancellation terms filed under the supplier's own 'What to bring' heading",
     "Phase 2: extend the line test to what_to_bring", "supplier"),
    ("SUPPLIER MISTAKE", "MEDIUM", 1, "680927",
     "Headings written with no content under them", "Correct behaviour; exclude from scoring",
     "supplier"),
    ("SUPPLIER MISTAKE", "LOW", 10,
     "509794, 203555, 249729, 330482, 279178, 444088, 397465, 319096, 171361, 324361",
     "Raw repeats itself -- the text shows twice on the page",
     "De-duplicate at render time; do NOT change extraction", "supplier"),
    ("SUPPLIER MISTAKE", "LOW", 4, "417608, 442752, 328897, 697755",
     "Near-duplicate, control character, 'n/a' placeholder, template keys pasted in",
     "Mostly nothing to do", "supplier"),
]


def load(fn):
    out = {}
    for line in (TEST_DIR / fn).open(encoding="utf-8"):
        d = json.loads(line)
        out[d["custom_id"].split("|")[0]] = json.loads(
            d["response"]["body"]["choices"][0]["message"]["content"]
        )
    return out


import re  # noqa: E402

# openpyxl refuses ASCII control characters outright (IllegalCharacterError).
# Supplier raw text carries them -- Bellingen Canoe Adventures' description has
# one mid-sentence. Strip control chars but keep newline and tab, which the
# wrapped cells rely on for readability.
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")


def xl(v):
    return ILLEGAL.sub("", v) if isinstance(v, str) else v


def main():
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    outs = load("v5_3_hard500_output.jsonl")
    scores = json.loads((TEST_DIR / "v5_3_hard500_scores.json").read_text(encoding="utf-8"))

    allp, perp = [], []
    for pid, fields in outs.items():
        sc = scores.get(pid, {})
        item = json.loads(Path(find_raw_file(pid)).read_text(encoding="utf-8"))["item"]
        sd = item.get("structured_description") or {}
        raw = strip_html(sd.get("description") or item.get("description") or "")
        name = item.get("name") or ""
        v, c = audit_verdict(pid)

        allp.append({
            "product_id": pid, "product_name": xl(name),
            "VERDICT": v, "COMMENT": xl(c),
            "fields_filled": sc.get("fields_filled"),
            "retention_pct": sc.get("retention_pct"),
            "duplicated": sc.get("duplicated_sentences"),
            "content_lost": len(sc.get("missing_sentences") or []),
            "headings_found": ", ".join(
                NICE.get(h, h) for h in (sc.get("headings_naming_a_column") or [])),
            "model_flags": xl(sc.get("model_flags", "")),
            "YOUR VERDICT": "", "YOUR NOTES": "",
        })
        rec = {"product_id": pid, "product_name": xl(name), "VERDICT": v,
               "RAW_DESCRIPTION": xl(raw)}
        for f in FIELDS:
            rec[NICE[f]] = xl(fields.get(f, ""))
        rec["model_flags"] = xl(fields.get("redo_flags", ""))
        perp.append(rec)

    SEV = {"CONTENT_LOSS": 0, "MISCLASS": 1, "SCHEMA": 2, "DUPLICATION": 2, "LABEL_LOSS": 3,
           "MINOR": 4, "SUPPLIER": 5, "OK": 9}
    df_all = pd.DataFrame(allp).assign(
        _s=lambda d: d["VERDICT"].map(SEV)).sort_values(["_s", "product_id"]).drop(columns="_s")
    df_iss = df_all[df_all["VERDICT"] != "OK"].copy()
    df_pp = pd.DataFrame(perp)
    df_mx = pd.DataFrame(MATRIX, columns=[
        "Group", "Severity", "Products", "Product IDs", "What is wrong",
        "Fix", "Whose"])

    with pd.ExcelWriter(OUT, engine="openpyxl") as w:
        df_mx.to_excel(w, index=False, sheet_name="Priority_Matrix")
        df_iss.to_excel(w, index=False, sheet_name="Issues_Only")
        df_all.to_excel(w, index=False, sheet_name="All_Products")
        df_pp.to_excel(w, index=False, sheet_name="Per_Product")

        head_fill = PatternFill("solid", fgColor="1F4E79")
        head_font = Font(color="FFFFFF", bold=True)
        widths = {"Group": 20, "Severity": 10, "What is wrong": 60, "Whose": 10, "Issue": 38, "Products": 9, "Product IDs": 46,
                  "Customer impact": 44, "Fix": 52, "Whose defect": 20,
                  "product_id": 12, "product_name": 34, "VERDICT": 14,
                  "COMMENT": 90, "RAW_DESCRIPTION": 70, "fields_filled": 9,
                  "retention_pct": 10, "duplicated": 10, "content_lost": 10}
        for sheet in ("Priority_Matrix", "Issues_Only", "All_Products", "Per_Product"):
            ws = w.sheets[sheet]
            ws.freeze_panes = "A2" if sheet == "Priority_Matrix" else "C2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                ws.column_dimensions[cell.column_letter].width = widths.get(
                    str(cell.value), 40)
            for r in ws.iter_rows(min_row=2):
                for c in r:
                    c.alignment = Alignment(vertical="top", wrap_text=True)

    print(f"wrote {OUT}")
    print(f"  Priority_Matrix : {len(df_mx)} rows")
    print(f"  Issues_Only     : {len(df_iss)} products")
    print(f"  All_Products    : {len(df_all)} products")
    print(f"  Per_Product     : {len(df_pp)} x {len(df_pp.columns)} cols")


if __name__ == "__main__":
    main()
