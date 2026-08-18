"""
Build luna100_manager_review.xlsx -- a workbook a non-technical reader can
check without help.

Design decisions, and why:

  ONE ROW PER PRODUCT, not per field. 100 rows a manager can scroll, rather
  than 2,800 rows they will close.

  RAW TEXT IMMEDIATELY BEFORE THE EXTRACTED FIELDS. A spot-check is then
  left-to-right on a single row -- no switching sheets, no lookups.

  PLAIN-ENGLISH COLUMN NAMES. "repeated_sentences", not "dup_sentences". The
  technical definition lives in the Summary sheet.

  EMPTY FIELDS ARE LEFT EMPTY and the Summary says so. Most tours have no
  itinerary and no cancellation policy; a blank is the correct answer and a
  manager scrolling past empty columns must not read that as failure.

Sheets:
  Summary          headline numbers + what they mean + what this does not prove
  Products         the 100 rows, with review columns
  Per_Field_Fill   how often each field was populated, with an example
  Needs_Attention  only the rows worth a second look

Usage:
    python build_luna100_workbook.py
"""
import sys
import re
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from screen_model_comparison import DESC_FIELDS, BOOKING_FIELDS, ALL_FIELDS

sys.stdout.reconfigure(encoding="utf-8")

TEST_DIR = Path(__file__).resolve().parent
SCREEN = TEST_DIR / "luna100_screen_results.json"
PRODUCTS = TEST_DIR / "luna100_products.json"
V500 = TEST_DIR / "v500_products.xlsx"
OUT = TEST_DIR / "luna100_manager_review.xlsx"

MODEL = "gpt-5.6-luna"
CELL_LIMIT = 32000
ILLEGAL_XL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
RAW_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
REVIEW_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
FIELD_FILL = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
AMBER = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
RED = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

# plain-English headers for the extracted fields
NICE = {
    "redo_desc_about": "About",
    "redo_desc_highlights": "Highlights",
    "redo_desc_what_included": "What's Included",
    "redo_desc_what_excluded": "What's NOT Included",
    "redo_desc_itinerary": "Itinerary",
    "redo_desc_what_to_bring": "What to Bring",
    "redo_desc_duration_text": "Duration",
    "redo_desc_requirements": "Requirements",
    "redo_desc_cancellation": "Cancellation Policy",
    "redo_desc_check_in": "Check-in",
    "redo_min_age": "Min Age",
    "redo_max_age": "Max Age",
    "redo_group_size": "Group Size",
    "redo_meeting_point": "Meeting Point",
    "redo_desc_other": "Other (description)",
    "redo_booking_what_to_bring": "Booking: What to Bring",
    "redo_booking_what_not_to_bring": "Booking: What NOT to Bring",
    "redo_booking_inclusions": "Booking: Inclusions",
    "redo_booking_location": "Booking: Location",
    "redo_booking_check_in": "Booking: Check-in",
    "redo_booking_departure_info": "Booking: Departure Info",
    "redo_booking_itinerary": "Booking: Itinerary",
    "redo_booking_important_info": "Booking: Important Info",
    "redo_booking_cancellation": "Booking: Cancellation",
    "redo_booking_faqs": "Booking: FAQs",
    "redo_booking_before_arrival": "Booking: Before Arrival",
    "redo_booking_contact": "Booking: Contact",
    "redo_booking_other": "Other (booking)",
}


def clean(t):
    s = "" if t is None else str(t)
    s = ILLEGAL_XL_RE.sub("", s)
    return s[:CELL_LIMIT] if len(s) > CELL_LIMIT else s


def style(ws, widths=None, wrap=None, fills=None, freeze="C2"):
    ws.freeze_panes = freeze
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = (fills or {}).get(c, HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in (widths or {}).items():
        if c <= ws.max_column:
            ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = (Alignment(wrap_text=True, vertical="top")
                              if wrap and cell.column in wrap
                              else Alignment(vertical="top"))


def main():
    screen = json.loads(SCREEN.read_text(encoding="utf-8"))[MODEL]
    meta = json.loads(PRODUCTS.read_text(encoding="utf-8"))
    ids = meta["product_ids"]

    v500 = pd.read_excel(V500, sheet_name="Summary")
    v500["product_id"] = v500["product_id"].astype(str)
    supplier = dict(zip(v500.product_id, v500.supplier_alias))

    # ---------- Products ----------
    rows = []
    for pid in ids:
        d = screen[pid]
        fv = d["field_values"]
        rows.append({
            "product_id": pid,
            "supplier": supplier.get(pid, ""),
            "words_in_source": d["input_words"],
            "fields_filled": d["fields_filled"],
            "text_kept_pct": round(d["word_coverage_pct"], 1),
            "size_ratio": d["word_ratio"],
            "repeated_sentences": d["dup_sentences"],
            "raw_description": clean(d["raw_desc"]),
            "raw_booking_notes": clean(d["raw_booking"]),
            **{NICE[f]: clean(fv.get(f, "")) for f in ALL_FIELDS},
            "pass/fail": "",
            "comment": "",
        })
    prod_df = pd.DataFrame(rows)

    # ---------- Per_Field_Fill ----------
    fill_rows = []
    for f in ALL_FIELDS:
        filled = [pid for pid in ids
                  if str(screen[pid]["field_values"].get(f) or "").strip()]
        example = ""
        if filled:
            example = " ".join(
                str(screen[filled[0]]["field_values"][f]).split())[:300]
        fill_rows.append({
            "field": NICE[f],
            "products_filled": len(filled),
            "fill_rate_pct": round(100 * len(filled) / len(ids), 1),
            "side": "description" if f in DESC_FIELDS else "booking notes",
            "example_value": clean(example),
            "example_from_product": filled[0] if filled else "",
        })
    fill_df = pd.DataFrame(fill_rows).sort_values("products_filled", ascending=False)

    # ---------- Itinerary check ----------
    # A known weakness of this model, found by hand-checking: it fills the
    # Itinerary field with ordinary narrative that merely mentions events in
    # order. A real itinerary needs a STRUCTURAL signal -- a clock time, or
    # "Day 2" / "Stop 3". Ordering words alone ("then", "next") are not enough,
    # because every story has a "then". Counted here so the scale is visible
    # rather than anecdotal.
    CLOCK = re.compile(r"\d{1,2}[:.]\d{2}\s*(am|pm)?|\b\d{1,2}\s*(am|pm)\b", re.I)
    STEP = re.compile(r"\b(day|stop|leg)\s*\d", re.I)
    itin_rows = []
    for pid in ids:
        for f in ("redo_desc_itinerary", "redo_booking_itinerary"):
            v = str(screen[pid]["field_values"].get(f) or "").strip()
            if not v:
                continue
            has_signal = bool(CLOCK.search(v) or STEP.search(v))
            itin_rows.append({
                "product_id": pid,
                "supplier": supplier.get(pid, ""),
                "field": NICE[f],
                "verdict": "looks like a real itinerary" if has_signal
                           else "QUESTIONABLE - no time or step number",
                "extracted_text": clean(v),
                "raw_description": clean(screen[pid]["raw_desc"]),
                "pass/fail": "", "comment": "",
            })
    itin_df = pd.DataFrame(itin_rows)
    n_weak = int((itin_df.verdict.str.startswith("QUESTIONABLE")).sum()) if len(itin_df) else 0

    # ---------- Needs_Attention ----------
    att = []
    for pid in ids:
        d = screen[pid]
        reasons = []
        if d["dup_sentences"]:
            reasons.append(f"{d['dup_sentences']} repeated sentence(s)")
        if d["untraceable_fields"]:
            reasons.append(f"{d['untraceable_fields']} field(s) not traceable to source")
        if d["word_ratio"] > 1.05:
            reasons.append(f"output {d['word_ratio']:.2f}x source (repeating)")
        elif d["word_ratio"] < 0.85:
            reasons.append(f"output {d['word_ratio']:.2f}x source (dropping text)")
        if d["markdown_fields"]:
            reasons.append(f"{d['markdown_fields']} field(s) contain markdown")
        if reasons:
            att.append({
                "product_id": pid, "supplier": supplier.get(pid, ""),
                "why_flagged": "; ".join(reasons),
                "text_kept_pct": round(d["word_coverage_pct"], 1),
                "size_ratio": d["word_ratio"],
                "repeated_sentences": d["dup_sentences"],
                "raw_description": clean(d["raw_desc"]),
                "pass/fail": "", "comment": "",
            })
    att_df = pd.DataFrame(att)

    # ---------- Summary ----------
    n = len(ids)
    vals = [screen[p] for p in ids]
    raw_tot = sum(v["input_words"] for v in vals)
    emit_tot = sum(v["words_emitted"] for v in vals)
    summary = [
        ("WHAT THIS IS", ""),
        ("Model tested", MODEL),
        ("Products", f"{n}, randomly sampled to match the real mix of short and "
                     f"long descriptions"),
        ("Cost to process all 23,034 products with this model", "$49"),
        ("", ""),
        ("HEADLINE NUMBERS", ""),
        ("Text kept", f"{sum(v['word_coverage_pct'] for v in vals) / n:.1f}% "
                      f"of the supplier's words appear somewhere in the output"),
        ("Size ratio", f"{emit_tot / raw_tot:.3f}  "
                       f"(1.00 = wrote exactly as much as the supplier did)"),
        ("Repeated sentences", f"{sum(v['dup_sentences'] for v in vals)} across all "
                               f"{n} products"),
        ("Invented content", f"{sum(v['untraceable_fields'] for v in vals)} field(s) "
                             f"could not be traced back to the supplier's text"),
        ("Fields filled", f"{sum(v['fields_filled'] for v in vals) / n:.1f} of 28 "
                          f"on average"),
        ("Products needing a look", f"{len(att_df)} of {n} (see Needs_Attention)"),
        ("", ""),
        ("THE ONE REAL PROBLEM FOUND", ""),
        ("Itinerary field is unreliable",
         f"{len(itin_df)} products got an Itinerary filled in. {n_weak} of those "
         f"({100 * n_weak / len(itin_df):.0f}%) contain ordinary description "
         f"text with no times and no numbered stops - so they are probably not "
         f"itineraries at all. See the Itinerary_Check sheet."),
        ("Why it happens", "The instructions currently accept words like 'then' "
                           "and 'next' as evidence of an itinerary. Every story "
                           "contains a 'then'. The fix is to require a clock time "
                           "or a numbered step, and it is a change to the "
                           "instructions, not to the AI model."),
        ("", ""),
        ("HOW TO READ THIS", ""),
        ("Products sheet", "One row per product. The supplier's original text is in "
                           "the two yellow columns; everything to the right is what "
                           "the AI pulled out of it. Read left to right to check a "
                           "product."),
        ("Empty fields are CORRECT", "Most tours have no itinerary and no "
                                     "cancellation policy. A blank box means the "
                                     "supplier never wrote one. An empty field is a "
                                     "better answer than an invented one."),
        ("Size ratio", "Words written divided by words supplied. Near 1.00 is right. "
                       "Above 1.1 means the AI repeated itself; below 0.9 means it "
                       "dropped content."),
        ("Repeated sentences", "The same sentence placed in two different boxes, "
                               "which would show the same paragraph twice on the "
                               "website. Counted within the description and within "
                               "the booking notes separately - the same fact "
                               "appearing in both is correct, not a fault."),
        ("", ""),
        ("WHAT THIS DOES NOT PROVE", ""),
        ("Placement is not measured", "These numbers show whether the supplier's "
                                      "text SURVIVED. They do not show whether each "
                                      "piece landed in the RIGHT box. That needs a "
                                      "person to read it."),
        ("Sample size", f"{n} products out of 11,236 Fareharbor products."),
        ("No independent judge", "No second AI reviewed this output. Earlier in this "
                                 "project, human review caught problems that two "
                                 "automated checks both missed."),
    ]
    sum_df = pd.DataFrame(summary, columns=["", " "])

    with pd.ExcelWriter(OUT, engine="openpyxl") as w:
        sum_df.to_excel(w, sheet_name="Summary", index=False)
        prod_df.to_excel(w, sheet_name="Products", index=False)
        fill_df.to_excel(w, sheet_name="Per_Field_Fill", index=False)
        if len(itin_df):
            itin_df.to_excel(w, sheet_name="Itinerary_Check", index=False)
        if len(att_df):
            att_df.to_excel(w, sheet_name="Needs_Attention", index=False)

        ws = w.sheets["Summary"]
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 100
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            a = row[0]
            if a.value and not str(row[1].value or "").strip():
                a.font = Font(bold=True, size=12)
                a.fill = HEADER_FILL

        cols = list(prod_df.columns)
        i_raw1 = cols.index("raw_description") + 1
        i_raw2 = cols.index("raw_booking_notes") + 1
        i_pf = cols.index("pass/fail") + 1
        i_cm = cols.index("comment") + 1
        fills = {i_raw1: RAW_FILL, i_raw2: RAW_FILL,
                 i_pf: REVIEW_FILL, i_cm: REVIEW_FILL}
        for c in range(i_raw2 + 1, i_pf):
            fills[c] = FIELD_FILL
        widths = {1: 12, 2: 22, 3: 10, 4: 9, 5: 11, 6: 10, 7: 12,
                  i_raw1: 70, i_raw2: 70, i_pf: 10, i_cm: 30}
        for c in range(i_raw2 + 1, i_pf):
            widths[c] = 42
        style(w.sheets["Products"], widths,
              wrap=set(range(i_raw1, i_cm + 1)), fills=fills)

        ws = w.sheets["Products"]
        i_ratio = cols.index("size_ratio") + 1
        i_dup = cols.index("repeated_sentences") + 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, i_ratio).value
            if isinstance(v, (int, float)):
                ws.cell(r, i_ratio).fill = (
                    GREEN if 0.85 <= v <= 1.05 else AMBER if v <= 1.15 else RED)
            d = ws.cell(r, i_dup).value
            if isinstance(d, (int, float)):
                ws.cell(r, i_dup).fill = (
                    GREEN if d == 0 else AMBER if d <= 3 else RED)

        style(w.sheets["Per_Field_Fill"],
              {1: 28, 2: 16, 3: 14, 4: 16, 5: 80, 6: 20}, wrap={5}, freeze="A2")

        if len(itin_df):
            ws = w.sheets["Itinerary_Check"]
            style(ws, {1: 12, 2: 22, 3: 18, 4: 34, 5: 70, 6: 70, 7: 10, 8: 30},
                  wrap={5, 6, 8}, freeze="C2")
            for r in range(2, ws.max_row + 1):
                v = str(ws.cell(r, 4).value or "")
                ws.cell(r, 4).fill = RED if v.startswith("QUESTIONABLE") else GREEN
        if len(att_df):
            style(w.sheets["Needs_Attention"],
                  {1: 12, 2: 22, 3: 55, 4: 13, 5: 11, 6: 18, 7: 70, 8: 10, 9: 30},
                  wrap={3, 7, 9}, freeze="C2")

    print("=" * 74)
    print("MANAGER REVIEW WORKBOOK")
    print("=" * 74)
    print(f"  products            : {n}")
    print(f"  text kept           : {sum(v['word_coverage_pct'] for v in vals) / n:.1f}%")
    print(f"  size ratio          : {emit_tot / raw_tot:.3f}")
    print(f"  repeated sentences  : {sum(v['dup_sentences'] for v in vals)}")
    print(f"  invented fields     : {sum(v['untraceable_fields'] for v in vals)}")
    print(f"  avg fields filled   : {sum(v['fields_filled'] for v in vals) / n:.1f} of 28")
    print(f"  needs attention     : {len(att_df)} products")
    print(f"  itineraries filled  : {len(itin_df)} ({n_weak} questionable)")
    print(f"\nWrote {OUT.name}")


if __name__ == "__main__":
    main()
