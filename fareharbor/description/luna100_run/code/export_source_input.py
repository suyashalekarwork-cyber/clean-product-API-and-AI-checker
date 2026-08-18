"""
Export the SOURCE INPUT for the 100 products -- the supplier's own text.

This is what goes INTO the extraction. Nothing here is model output.

    python export_source_input.py

Writes to data/:
    luna100_source_input.json   the input, keyed by product id
    luna100_source_input.csv    the same, one row per product
    luna100_source_input.xlsx   the same, readable, header frozen

Each product carries the two raw fields the prompt is given:

    raw_description     the supplier's description text  (all 100 products)
    raw_booking_notes   the supplier's booking notes     (66 products; 34 have none)

Anyone can take this file, run it through any prompt or any model, and compare
the result against output/ or extracted/ in this folder.
"""
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
MODEL = "gpt-5.6-luna"
CELL_LIMIT = 32000
ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(t):
    return ILLEGAL.sub("", "" if t is None else str(t))[:CELL_LIMIT]


def main():
    screen = json.loads((DATA / "luna100_screen_results.json")
                        .read_text(encoding="utf-8"))[MODEL]
    meta = json.loads((DATA / "luna100_products.json").read_text(encoding="utf-8"))
    ids = meta["product_ids"]

    supplier = {}
    try:
        v = pd.read_excel(DATA / "v500_products.xlsx", sheet_name="Summary")
        v["product_id"] = v["product_id"].astype(str)
        supplier = dict(zip(v.product_id, v.supplier_alias))
    except Exception:
        pass

    src, rows = {}, []
    for pid in ids:
        desc = screen[pid].get("raw_desc") or ""
        book = screen[pid].get("raw_booking") or ""
        src[pid] = {
            "product_id": pid,
            "supplier": supplier.get(pid, ""),
            "source": "Fareharbor",
            "raw_description": desc,
            "raw_booking_notes": book,
        }
        rows.append({
            "product_id": pid,
            "supplier": supplier.get(pid, ""),
            "has_description": "yes" if desc.strip() else "no",
            "has_booking_notes": "yes" if book.strip() else "no",
            "description_words": len(desc.split()),
            "booking_words": len(book.split()),
            "raw_description": clean(desc),
            "raw_booking_notes": clean(book),
        })

    df = pd.DataFrame(rows)
    (DATA / "luna100_source_input.json").write_text(
        json.dumps(src, indent=2, ensure_ascii=False), encoding="utf-8",
        newline="\n")
    df.to_csv(DATA / "luna100_source_input.csv", index=False,
              encoding="utf-8-sig")
    with pd.ExcelWriter(DATA / "luna100_source_input.xlsx",
                        engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Source_Input", index=False)
        ws = w.sheets["Source_Input"]
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions
        raw = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
        hdr = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = raw if c in (7, 8) else hdr
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(c)].width = (
                12 if c == 1 else 22 if c == 2 else 14 if c < 7 else 80)
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    nd = int((df.has_description == "yes").sum())
    nb = int((df.has_booking_notes == "yes").sum())
    print(f"Wrote data/luna100_source_input.[json|csv|xlsx]")
    print(f"  {len(df)} products")
    print(f"  {nd} with description text, {nb} with booking notes "
          f"({nd + nb} prompt requests per run)")
    print(f"  description: {df.description_words.min()}-"
          f"{df.description_words.max()} words, median "
          f"{int(df.description_words.median())}")


if __name__ == "__main__":
    main()
