"""
Export the extracted data from a run into readable files.

The files in output/ are raw Batch API replies -- one JSON envelope per line,
with the extracted fields buried inside a string. This unpacks them.

    python export_extracted.py            # every run in output/
    python export_extracted.py --run luna100_v4_8_3_output.jsonl

Writes to extracted/, per run:

    luna100_<version>_extracted.csv    one row per product, every field, plus
                                       the raw supplier text
    luna100_<version>_extracted.json   the same data keyed by product id
    luna100_<version>_extracted.xlsx   spreadsheet, raw text beside the fields,
                                       frozen header, filters on

Description and booking sides are merged onto one row per product, so a product
appears once with all 29 fields.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "output"
DEST = ROOT / "extracted"
MODEL = "gpt-5.6-luna"

# plain-English column names, in the order a reader wants them
NICE = [
    ("redo_desc_about", "About"),
    ("redo_desc_highlights", "Highlights"),
    ("redo_desc_what_included", "What's Included"),
    ("redo_desc_what_excluded", "What's NOT Included"),
    ("redo_desc_itinerary", "Itinerary"),
    ("redo_desc_faqs", "FAQs"),
    ("redo_desc_what_to_bring", "What to Bring"),
    ("redo_desc_duration_text", "Duration"),
    ("redo_desc_requirements", "Requirements"),
    ("redo_desc_cancellation", "Cancellation Policy"),
    ("redo_desc_check_in", "Check-in"),
    ("redo_min_age", "Min Age"),
    ("redo_max_age", "Max Age"),
    ("redo_group_size", "Group Size"),
    ("redo_meeting_point", "Meeting Point"),
    ("redo_desc_other", "Other (description)"),
    ("redo_booking_what_to_bring", "Booking: What to Bring"),
    ("redo_booking_what_not_to_bring", "Booking: What NOT to Bring"),
    ("redo_booking_inclusions", "Booking: Inclusions"),
    ("redo_booking_location", "Booking: Location"),
    ("redo_booking_check_in", "Booking: Check-in"),
    ("redo_booking_departure_info", "Booking: Departure Info"),
    ("redo_booking_itinerary", "Booking: Itinerary"),
    ("redo_booking_important_info", "Booking: Important Info"),
    ("redo_booking_cancellation", "Booking: Cancellation"),
    ("redo_booking_faqs", "Booking: FAQs"),
    ("redo_booking_before_arrival", "Booking: Before Arrival"),
    ("redo_booking_contact", "Booking: Contact"),
    ("redo_booking_other", "Other (booking)"),
]
CELL_LIMIT = 32000
ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(t):
    return ILLEGAL.sub("", "" if t is None else str(t))[:CELL_LIMIT]


def parse(path):
    """product_id -> merged dict of both sides. Repairs the model's JSON defect."""
    out, repaired, failed = {}, [], []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        cid = r.get("custom_id", "")
        if "|" not in cid:
            continue
        pid = cid.split("|")[0]
        try:
            t = r["response"]["body"]["choices"][0]["message"]["content"].strip()
            t = re.sub(r"^```(?:json)?\s*|\s*```$", "", t).strip()
        except Exception:
            failed.append(cid)
            continue
        try:
            d = json.loads(t)
        except json.JSONDecodeError:
            # the model intermittently closes its JSON with a stray `,"`
            try:
                d = json.loads(re.sub(r',\s*"\s*\}\s*$', "}", t))
                repaired.append(cid)
            except json.JSONDecodeError:
                failed.append(cid)
                continue
        out.setdefault(pid, {}).update(d)
    return out, repaired, failed


def style(ws):
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    raw = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    hdr = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
    fld = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = raw if c in (3, 4) else (fld if c > 4 else hdr)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(c)].width = (
            12 if c == 1 else 22 if c == 2 else 60 if c in (3, 4) else 40)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def export(path, screen, ids, supplier):
    tag = re.search(r"luna100_(.+?)_output", path.name).group(1)
    fields, repaired, failed = parse(path)

    rows, jrows = [], {}
    for pid in ids:
        d = fields.get(pid, {})
        row = {
            "product_id": pid,
            "supplier": supplier.get(pid, ""),
            "RAW description (source)": clean(screen[pid].get("raw_desc")),
            "RAW booking notes (source)": clean(screen[pid].get("raw_booking")),
        }
        for key, label in NICE:
            row[label] = clean(d.get(key, ""))
        rows.append(row)
        jrows[pid] = {
            "supplier": supplier.get(pid, ""),
            "raw_description": screen[pid].get("raw_desc") or "",
            "raw_booking_notes": screen[pid].get("raw_booking") or "",
            "extracted": {k: (d.get(k) or "") for k, _ in NICE},
        }

    df = pd.DataFrame(rows)
    DEST.mkdir(exist_ok=True)
    base = DEST / f"luna100_{tag}_extracted"
    df.to_csv(base.with_suffix(".csv"), index=False, encoding="utf-8-sig")
    base.with_suffix(".json").write_text(
        json.dumps(jrows, indent=2, ensure_ascii=False), encoding="utf-8",
        newline="\n")
    with pd.ExcelWriter(base.with_suffix(".xlsx"), engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Extracted", index=False)
        style(w.sheets["Extracted"])

    filled = {label: int((df[label].str.strip() != "").sum()) for _, label in NICE}
    print(f"\n{path.name}  ->  extracted/luna100_{tag}_extracted.[csv|json|xlsx]")
    print(f"   {len(df)} products x {len(NICE)} fields"
          + (f"   |  JSON repaired: {len(repaired)}" if repaired else "")
          + (f"   |  FAILED: {failed}" if failed else ""))
    top = ", ".join(f"{k} {v}" for k, v in
                    sorted(filled.items(), key=lambda x: -x[1])[:6])
    print(f"   most filled: {top}")
    return tag, filled


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", help="a single file in output/ (default: all)")
    args = ap.parse_args()

    screen = json.loads((DATA / "luna100_screen_results.json")
                        .read_text(encoding="utf-8"))[MODEL]
    ids = json.loads((DATA / "luna100_products.json")
                     .read_text(encoding="utf-8"))["product_ids"]
    supplier = {p: screen[p].get("supplier", "") for p in ids}
    try:
        v500 = pd.read_excel(DATA / "v500_products.xlsx", sheet_name="Summary")
        v500["product_id"] = v500["product_id"].astype(str)
        supplier = dict(zip(v500.product_id, v500.supplier_alias))
    except Exception:
        pass

    runs = ([OUT / args.run] if args.run
            else sorted(OUT.glob("luna100_*_output.jsonl")))
    if not runs or not runs[0].exists():
        raise SystemExit("no run output found in output/")

    summary = {}
    for p in runs:
        tag, filled = export(p, screen, ids, supplier)
        summary[tag] = filled

    if len(summary) > 1:
        print("\nFields filled, per version:")
        labels = [l for _, l in NICE]
        hdr = "  " + "field".ljust(28) + "".join(t.rjust(10) for t in summary)
        print(hdr)
        print("  " + "-" * (len(hdr) - 2))
        for l in labels:
            vals = [summary[t][l] for t in summary]
            if any(vals):
                print("  " + l.ljust(28)
                      + "".join(str(v).rjust(10) for v in vals))


if __name__ == "__main__":
    main()
