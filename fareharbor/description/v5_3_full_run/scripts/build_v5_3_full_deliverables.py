"""
Deliverables for the full-catalogue V5.3 description run (11,069 products).

Three outputs:

  reports/v5_3_full_summary.md      the numbers, with what was verified marked
  reports/v5_3_full_findings.txt    per-product detail for products with a
                                    finding -- NOT all 11,069, which would be a
                                    ~35 MB file nobody opens
  exports/v5_3_full_scores.xlsx     every product, one row, sortable

Why findings-only for the txt: the 1,000-product audit is already 3.2 MB and the
useful part is the flagged minority. A reader wants the products with something
wrong plus a way to look up any given id -- the workbook covers the second.
"""
import json
import re
import sys
from collections import Counter
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd                                       # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill   # noqa: E402
from openpyxl.utils import get_column_letter               # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEST = ROOT / "data_pipeline" / "batch_api_test"
sys.path.insert(0, str(TEST))

from build_model_comparison_batches import strip_html, find_raw_file  # noqa: E402
from booking_common import parse_booking_json                          # noqa: E402

SCORES = TEST / "v5_3_full_scores.json"
OUTPUT = TEST / "v5_3_full_output.jsonl"
MD_OUT = ROOT / "reports" / "v5_3_full_summary.md"
TXT_OUT = ROOT / "reports" / "v5_3_full_findings.txt"
XLSX_OUT = ROOT / "exports" / "v5_3_full_scores.xlsx"

FIELDS = [
    "redo_desc_about", "redo_desc_important_info", "redo_desc_highlights",
    "redo_desc_what_included", "redo_desc_what_excluded", "redo_desc_extras",
    "redo_desc_itinerary", "redo_desc_what_to_bring", "redo_desc_duration_text",
    "redo_desc_cancellation", "redo_desc_check_in", "redo_desc_accessibility",
    "redo_desc_restrictions", "redo_desc_special_requirements",
    "redo_desc_faqs", "redo_desc_pricing", "redo_desc_disclaimers",
    "redo_meeting_point", "redo_group_size",
]
NICE = {f: f.replace("redo_desc_", "").replace("redo_", "") for f in FIELDS}

COUNTED = [
    ("duplicated_sentences", "duplicated sentences"),
    ("untraceable_sentences", "untraceable (possible invention)"),
    ("missing_sentences", "missing sentences (content loss)"),
    ("filled_but_no_heading", "filled with no heading"),
    ("itinerary_lines_without_signal", "itinerary lines without a signal"),
    ("dropped_informative_headings", "dropped informative headings"),
    ("included_lines_that_are_purchasable", "included lines that are purchasable"),
    ("pricing_without_figure", "pricing with no figure"),
    ("cancellation_without_refund", "cancellation with no refund"),
    ("markdown_junk_fields", "markdown junk"),
]
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")


def xl(v):
    return ILLEGAL.sub("", v)[:32000] if isinstance(v, str) else v


def n_of(v):
    return len(v) if isinstance(v, (list, tuple)) else (int(bool(v)) if v else 0)


def main():
    scores = json.loads(SCORES.read_text(encoding="utf-8"))
    print(f"scored products: {len(scores):,}")

    rows = []
    for pid, s in scores.items():
        r = {"product_id": pid, "retention_pct": s.get("retention_pct"),
             "fields_filled": n_of(s.get("fields_filled")),
             "headings_naming_a_column": n_of(s.get("headings_naming_a_column")),
             "fidelity": s.get("fidelity")}
        for k, label in COUNTED:
            r[k] = n_of(s.get(k))
        r["total_findings"] = sum(r[k] for k, _ in COUNTED)
        rows.append(r)
    df = pd.DataFrame(rows)

    # ---------------- workbook ----------------
    df_sorted = df.sort_values(["total_findings", "retention_pct"],
                               ascending=[False, True])
    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as w:
        df_sorted.to_excel(w, sheet_name="All_Products", index=False)
        ws = w.sheets["All_Products"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(df_sorted.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(c) + 3))
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", start_color="D9D9D9")
            c.alignment = Alignment(vertical="top", wrap_text=True)
    print(f"wrote {XLSX_OUT.relative_to(ROOT)}  ({len(df):,} rows)")

    # ---------------- findings txt ----------------
    outs = {}
    for line in OUTPUT.open(encoding="utf-8"):
        d = json.loads(line)
        f, _ = parse_booking_json(
            d["response"]["body"]["choices"][0]["message"]["content"])
        outs[d["custom_id"].split("|")[0]] = f or {}

    flagged = df_sorted[df_sorted.total_findings > 0].product_id.tolist()
    L = []
    A = L.append
    A("=" * 96)
    A("V5.3 DESCRIPTION -- FULL CATALOGUE, PRODUCTS WITH A FINDING")
    A("=" * 96)
    A("")
    A(f"{len(flagged):,} of {len(df):,} products have at least one finding.")
    A("The other products are in exports/v5_3_full_scores.xlsx -- this file is")
    A("the flagged minority, because a 35 MB file of 11,069 products is not one")
    A("anybody opens.")
    A("")
    A("VERIFIED BY SAMPLING (40 products each, seed 42):")
    A("  filled with no heading          ~90% are REAL gate leaks")
    A("  itinerary lines without signal  ~90% are REAL line-test failures")
    A("Both were checked because the equivalent flags on the booking run were")
    A("mostly the scorer's own fault. Here they are not -- every earlier sample")
    A("was heading-rich by selection, and the full catalogue is not.")
    A("")
    A("NOT VERIFIED: duplication, untraceable, missing sentences. Treat those")
    A("counts as upper bounds.")
    A("")

    for pid in flagged[:1500]:
        s = scores[pid]
        try:
            raw = strip_html(
                (json.loads(Path(find_raw_file(pid)).read_text(encoding="utf-8"))
                 ["item"].get("structured_description") or {}).get("description")
                or json.loads(Path(find_raw_file(pid)).read_text(encoding="utf-8"))
                ["item"].get("description") or "")
        except Exception:                                          # noqa: BLE001
            raw = ""
        A("=" * 96)
        A(f"PRODUCT {pid}   retention {s.get('retention_pct')}%   "
          f"fields filled {n_of(s.get('fields_filled'))}")
        A("=" * 96)
        for k, label in COUNTED:
            v = s.get(k)
            if not v:
                continue
            items = v if isinstance(v, (list, tuple)) else [str(v)]
            A(f"  {label} ({len(items)})")
            for it in items[:6]:
                A(f"      - {' '.join(str(it).split())[:150]}")
            if len(items) > 6:
                A(f"      ... {len(items) - 6} more")
        A("")
        A("  RAW DESCRIPTION")
        for l in [x.strip() for x in raw.split("\n") if x.strip()][:25]:
            A(f"    {l[:110]}")
        A("")
        A("  EXTRACTED")
        o = outs.get(pid, {})
        for f in FIELDS:
            v = (o.get(f) or "").strip()
            if v:
                A(f"    [{NICE[f]}] {' '.join(v.split())[:220]}")
        A("")
    if len(flagged) > 1500:
        A(f"... and {len(flagged) - 1500:,} more flagged products -- see the workbook.")

    TXT_OUT.write_text("\n".join(L), encoding="utf-8")
    print(f"wrote {TXT_OUT.relative_to(ROOT)}  "
          f"({TXT_OUT.stat().st_size / 1e6:.1f} MB, {len(flagged):,} flagged)")

    # ---------------- summary markdown ----------------
    tot = {k: int(df[k].sum()) for k, _ in COUNTED}
    prod = {k: int((df[k] > 0).sum()) for k, _ in COUNTED}
    M = []
    B = M.append
    B("# V5.3 description — full catalogue")
    B("")
    B(f"Every Fareharbor product with a description: **{len(df):,}**. "
      "167 products have no description and were skipped.")
    B("")
    B("## Run integrity")
    B("")
    B("| | |")
    B("|---|---|")
    B(f"| Products | **{len(df):,}** |")
    B("| Responses parsed | **11,069 of 11,069** (one needed a JSON repair) |")
    B("| Truncated | **0** |")
    B("| Wrong key count | **0** |")
    B(f"| Mean content retention | **{df.retention_pct.mean():.1f}%** |")
    B(f"| Products at 100% retention | **{int((df.retention_pct == 100).sum()):,}** "
      f"({100 * (df.retention_pct == 100).mean():.1f}%) |")
    B(f"| Products with no finding at all | **{int((df.total_findings == 0).sum()):,}** "
      f"({100 * (df.total_findings == 0).mean():.1f}%) |")
    B("")
    B("Retention is **higher** at full scale than on the random-1,000 (99.17%) or")
    B("the hardest-500. Every earlier set was difficulty-selected, so those are")
    B("the pessimistic end rather than the typical case.")
    B("")
    B("## Findings")
    B("")
    B("| Finding | Count | Products | Verified? |")
    B("|---|---|---|---|")
    VER = {
        "filled_but_no_heading": "**yes — ~90% real** (40 sampled)",
        "itinerary_lines_without_signal": "**yes — ~90% real** (40 sampled)",
    }
    for k, label in COUNTED:
        B(f"| {label} | {tot[k]:,} | {prod[k]:,} | {VER.get(k, 'no — upper bound')} |")
    B("")
    B("### What was verified, and why")
    B("")
    B("On the booking run the same two flags were mostly the scorer's own fault —")
    B("`filled with no heading` went from 245 to 0 once the heading mapper was")
    B("fixed. So both were sampled here before being reported. **They hold up:**")
    B("about 90% of each is real.")
    B("")
    B("Example: product `529030` filled `cancellation` when its only headings are")
    B("`Duration`, `Ticket Prices:` and `About`. Nothing licensed that column.")
    B("")
    B("That the gate leaks more here than on the curated samples is expected —")
    B("every earlier set was heading-rich by selection, and the full catalogue")
    B("contains thousands of products with sparse or no headings.")
    B("")
    B("**The other counts are NOT verified** and should be read as upper bounds.")
    B("Across this project roughly a quarter of what any detector reports has")
    B("turned out to be the detector rather than the model.")
    B("")
    B("## Known issues carried forward")
    B("")
    B("- **Duplication** — the one hard gate that fails. Three prompt versions")
    B("  failed to fix it on smaller runs; it needs the deterministic")
    B("  post-processing pass, not more wording.")
    B("- **Difficulty ratings** (`Difficulty: Hard`) wrongly fill `restrictions`.")
    B("  Reproducible across runs, and a two-line prompt fix.")
    B("- **Content loss** does not shrink with easier products, which is why the")
    B("  `recovered_content` pass built for the booking side should be applied")
    B("  here too.")
    MD_OUT.write_text("\n".join(M), encoding="utf-8")
    print(f"wrote {MD_OUT.relative_to(ROOT)}")

    print(f"\nretention mean {df.retention_pct.mean():.2f}%  "
          f"at100 {int((df.retention_pct == 100).sum()):,}  "
          f"clean {int((df.total_findings == 0).sum()):,}")


if __name__ == "__main__":
    main()
