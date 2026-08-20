"""Excel workbook for exports/fareharbor_unified_v2.csv.

Three sheets:
  Products   the 11,231 rows, frozen header, autofilter, colour-banded by
             schema group so the 59 columns stay navigable
  Schema     one row per column - what it holds, fill count, fill %
  Read me    the rules a consumer must know before using the data

Reads the CSV as TEXT throughout. Numeric-looking ids and postcodes must not be
coerced - pandas turns a nullable int column into 2000.0 (CLAUDE.md, Data Quirks).
Writes to a NEW filename each run is NOT done; if the target is open in Excel
openpyxl raises PermissionError, which is reported rather than swallowed.
"""
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "exports", "fareharbor_unified_v2.csv")
OUT = os.path.join(ROOT, "exports", "fareharbor_unified_v2.xlsx")

# group -> (columns, header fill, band fill)
GROUPS = [
    ("Identity", ["compound_key", "product_id", "source", "product_name",
                  "product_headline"], "1F2A44", "EDEFF4"),
    ("Supplier", ["meta_supplier_id", "meta_supplier_name", "meta_operator_info"],
     "5B4B8A", "F0EDF6"),
    # Price withdrawn 2026-08-20 - it comes from a separate API. The group
    # keeps its colour and its remaining members rather than being deleted.
    ("Listing", ["product_duration", "product_duration_minutes",
                 "product_category", "product_tags"],
     "8A5300", "F8F1E6"),
    ("Location", ["location_street", "location_city", "location_state",
                  "location_country", "location_postcode", "location_latitude",
                  "location_longitude", "location_end"], "0B6E68", "E8F3F2"),
    ("Detail", ["detail_description", "detail_highlights", "detail_what_is_included",
                "detail_what_is_not_included", "detail_itinerary",
                "detail_important_info", "detail_booking_notes",
                "detail_meeting_point", "detail_check_in", "detail_departure_info",
                "detail_before_arrival", "detail_what_to_bring",
                "detail_accessibility", "detail_onboard_facilities",
                "detail_restrictions",
                "detail_special_requirements", "detail_health_safety",
                "detail_group_size", "detail_faqs", "detail_extras",
                "detail_disclaimers", "detail_cancellation_policy",
                "detail_cancellation_hours", "detail_operating_days",
                "detail_start_time", "detail_return_time", "detail_languages",
                "detail_pickup_available"], "1F4E9C", "EAF0F9"),
    ("Media", ["product_images", "product_videos"], "9C1F3C", "FAECF0"),
    ("Provenance", ["extractions_present"], "3A4358", "EFF1F5"),
]

HOLDS = {
    "compound_key": "product_id + source. The primary key.",
    "product_id": "Fareharbor `pk`, stored as text.",
    "source": "Always Fareharbor in this file.",
    "product_name": "Product title.",
    "product_headline": "Fareharbor `headline` — attribute chips, sometimes dated.",
    "meta_supplier_id": "Blank: the Details API carries no supplier data.",
    "meta_supplier_name": "Blank: same reason. Metadata API deliberately not used.",
    "meta_operator_info": "JSON. Only contact_text is filled, from booking notes.",
    "product_duration": "Supplier's own words. Never derived.",
    "product_duration_minutes": "Blank: Fareharbor gives prose only, and we derive nothing.",
    "product_category": "First tag.",
    "product_tags": "JSON array of all tags.",
    "location_street": "From locations[], not primary_location.",
    "location_city": "From locations[].",
    "location_state": "Fareharbor calls this `province`.",
    "location_country": "2-letter code.",
    "location_postcode": "Text — leading zeros preserved.",
    "location_latitude": "GPS.",
    "location_longitude": "GPS.",
    "location_end": "Blank: Fareharbor has no separate finish point.",
    "detail_description": "The un-headed remainder of the description.",
    "detail_what_to_bring": "Includes 'Do not bring:' items, merged in.",
    "detail_meeting_point": "structured_description.meeting_point + locations[].note.",
    "detail_cancellation_hours": "Hours notice. Fareharbor states hours natively.",
    "detail_onboard_facilities": "Blank in EVERY source. Holds the Figma Onboard Facilities section open; no API carries a facilities list.",
    "detail_pickup_available": "true / false.",
    "product_images": "JSON. The cover is flagged is_main — do NOT assume it is first.",
    "product_videos": "Blank: Fareharbor has no video field.",
    "extractions_present": "Which extractions ran: description, booking, or none.",
}

THIN = Side(style="thin", color="D5D9E3")


def main():
    df = pd.read_csv(SRC, low_memory=False, dtype=str).fillna("")
    cols = list(df.columns)
    print(f"{len(df):,} rows x {len(cols)} columns")

    gof, headfill, bandfill = {}, {}, {}
    for name, gcols, hx, bx in GROUPS:
        for c in gcols:
            gof[c], headfill[c], bandfill[c] = name, hx, bx

    missing = [c for c in cols if c not in gof]
    if missing:
        raise SystemExit(f"columns not assigned to a group: {missing}")

    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        # ---------------------------------------------------------- Products
        df.to_excel(xl, sheet_name="Products", index=False)
        ws = xl.book["Products"]
        ws.freeze_panes = "E2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
        ws.row_dimensions[1].height = 34

        WIDE = {"product_name": 42, "product_headline": 34, "detail_description": 52,
                "product_images": 34,
                "meta_operator_info": 34, "compound_key": 20}
        for i, c in enumerate(cols, 1):
            L = get_column_letter(i)
            cell = ws.cell(row=1, column=i)
            cell.value = c
            cell.font = Font(bold=True, size=9.5, color="FFFFFF", name="Consolas")
            cell.fill = PatternFill("solid", fgColor=headfill[c])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="1F2A44"))
            if c in WIDE:
                ws.column_dimensions[L].width = WIDE[c]
            elif c.startswith("detail_"):
                ws.column_dimensions[L].width = 30
            elif c.startswith("location_"):
                ws.column_dimensions[L].width = 16
            else:
                ws.column_dimensions[L].width = 18

        band = PatternFill("solid", fgColor="F7F8FB")
        for r in range(2, len(df) + 2):
            ws.row_dimensions[r].height = 15
            stripe = (r % 2 == 0)
            for i, c in enumerate(cols, 1):
                cell = ws.cell(row=r, column=i)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.font = Font(size=9.5, name="Consolas" if c in (
                    "compound_key", "product_id",
                    "product_images", "meta_operator_info") else "Calibri")
                cell.border = Border(bottom=THIN, right=THIN)
                if stripe:
                    cell.fill = band
        print("  Products sheet done")

        # ------------------------------------------------------------ Schema
        srows = []
        for name, gcols, _h, _b in GROUPS:
            for c in gcols:
                n = int((df[c].str.strip() != "").sum())
                srows.append({
                    "Group": name, "Column": c,
                    "What it holds": HOLDS.get(c, ""),
                    "Filled": n, "Fill %": round(n / len(df) * 100, 1),
                    "Empty": len(df) - n,
                })
        sdf = pd.DataFrame(srows)
        sdf.to_excel(xl, sheet_name="Schema", index=False)
        ss = xl.book["Schema"]
        ss.freeze_panes = "A2"
        ss.auto_filter.ref = f"A1:F{len(sdf) + 1}"
        for L, w in zip("ABCDEF", (13, 30, 62, 10, 10, 10)):
            ss.column_dimensions[L].width = w
        for i in range(1, 7):
            h = ss.cell(row=1, column=i)
            h.font = Font(bold=True, size=10, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor="1F2A44")
            h.alignment = Alignment(vertical="center")
        ss.row_dimensions[1].height = 24
        for r in range(2, len(sdf) + 2):
            pct = sdf.iloc[r - 2]["Fill %"]
            grp = sdf.iloc[r - 2]["Group"]
            for i in range(1, 7):
                cell = ss.cell(row=r, column=i)
                cell.font = Font(size=9.5,
                                 name="Consolas" if i == 2 else "Calibri",
                                 color="8A92AB" if pct == 0 else "000000")
                cell.border = Border(bottom=THIN)
                cell.alignment = Alignment(vertical="center", wrap_text=(i == 3))
            ss.cell(row=r, column=1).fill = PatternFill(
                "solid", fgColor=dict((g[0], g[3]) for g in GROUPS)[grp])
            f = ss.cell(row=r, column=5)
            f.number_format = "0.0"
            if pct == 0:
                f.fill = PatternFill("solid", fgColor="F1F2F6")
            elif pct >= 90:
                f.fill = PatternFill("solid", fgColor="D9EFD9")
            elif pct >= 40:
                f.fill = PatternFill("solid", fgColor="FFF4D6")
            else:
                f.fill = PatternFill("solid", fgColor="FBE4E4")
        print("  Schema sheet done")

        # ----------------------------------------------------------- Read me
        notes = [
            ("FAREHARBOR — UNIFIED SCHEMA v2", ""),
            ("", ""),
            (f"{len(df):,} products x {len(cols)} columns.",
             "Built from raw Details-API files plus the V5.3 description and "
             "V5.4 booking extractions. Nothing re-fetched, nothing re-extracted."),
            ("", ""),
            ("BEFORE YOU DISPLAY ANYTHING", ""),
            ("Strip the source tags.",
             "Every text column starts each block with [API], [DESCRIPTION] or "
             "[BOOKING NOTES]. Strip with the regex ^\\[[A-Z ]+\\]\\s* or agents "
             "will see the tags on the page. This is required, not optional."),
            ("[API] leads.",
             "It means the supplier's own dedicated field — the most trustworthy "
             "block. Extracted blocks follow it."),
            ("Empty means the supplier gave nothing.",
             "It is a fact about the feed, not a defect. Nothing is ever invented "
             "to fill a blank. Hide empty sections in the UI."),
            ("", ""),
            ("THINGS THAT WILL BITE", ""),
            ("Prices are already in dollars.",
             "The Fareharbor API returns CENTS. This file divides by 100. Do not "
             "divide again."),
            ("The main image is flagged, not first.",
             "In product_images, find the entry where is_main is true. On ~6% of "
             "products the cover is a photo that appears nowhere else in the "
             "gallery, and it has been added to the list."),
            ("Read ids and postcodes as text.",
             "Excel and pandas will otherwise turn 2000 into 2000.0 and drop "
             "leading zeros."),
            ("There is no price in this workbook.",
             "Pricing comes from a separate API and was withdrawn on 2026-08-20: "
             "product_price, product_currency, product_price_unit, "
             "product_price_tax_inclusive, product_price_options, "
             "product_min_quantity, product_max_quantity, detail_tax_percentage "
             "and detail_pricing_notes. Join on product_id. The values are still "
             "in the raw supplier files - nothing was destroyed."),
            ("", ""),
            ("COLUMNS THAT ARE ENTIRELY EMPTY, AND CORRECTLY SO", ""),
            ("Supplier name and id.",
             "Fareharbor's Details API carries no supplier data at all. The "
             "Metadata API is deliberately not used. Only meta_operator_info is "
             "filled, on 253 products, from contact details in booking notes."),
            ("Duration in minutes, videos, operating days, start and return "
             "time, end location.",
             "These are Rezdy, Livn and CustomLinc fields. Fareharbor supplies "
             "none of them. They exist so every source shares one schema."),
            ("Onboard facilities.",
             "Empty in EVERY source, not just Fareharbor. The Figma section "
             "exists but no supplier API carries a facilities list - the chip "
             "text appears only as prose inside descriptions, and generating a "
             "value from prose is the one thing this schema never does."),
            ("", ""),
            ("KNOWN GAPS", ""),
            ("167 products have no description; 2,992 have no booking notes; "
             "108 have neither.",
             "See the extractions_present column. Five further raw files are API "
             "error responses, not products — access forbidden or an invalid id — "
             "so the true catalogue is 11,231, not 11,236."),
            ("Contact details are under-captured.",
             "Only booking notes were extracted for contact. Roughly 3-4% of "
             "descriptions also carry a phone or email that was never asked for."),
            ("", ""),
            ("Full decision record: reports/ALL_SOURCES_FIELD_MAP.md", ""),
            ("Built by: scripts/build_fareharbor_v2.py", ""),
        ]
        ndf = pd.DataFrame(notes, columns=["", " "])
        ndf.to_excel(xl, sheet_name="Read me", index=False)
        ns = xl.book["Read me"]
        ns.column_dimensions["A"].width = 52
        ns.column_dimensions["B"].width = 92
        ns.sheet_view.showGridLines = False
        for r in range(1, len(notes) + 2):
            a, b = ns.cell(row=r, column=1), ns.cell(row=r, column=2)
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b.alignment = Alignment(vertical="top", wrap_text=True)
            a.font = Font(size=10.5)
            b.font = Font(size=10.5, color="3A4358")
            txt = str(a.value or "")
            if r == 2:
                a.font = Font(bold=True, size=17, color="1F2A44")
            elif txt.isupper() and txt:
                a.font = Font(bold=True, size=10, color="9C1F3C")
                ns.row_dimensions[r].height = 26
            elif txt:
                a.font = Font(bold=True, size=10.5, color="1F2A44")
        ns.cell(row=1, column=1).value = ""
        print("  Read me sheet done")

        xl.book.active = 0

    size = os.path.getsize(OUT) / 1e6
    print(f"\nwrote {OUT}  ({size:.1f} MB)")


if __name__ == "__main__":
    # Excel holds a write lock on an open workbook. Rather than fail the whole
    # rebuild, divert to a sibling filename and say so loudly - a run that
    # silently wrote nothing is worse than one that wrote somewhere else.
    if len(sys.argv) > 1:
        OUT = sys.argv[1]
    try:
        main()
    except PermissionError:
        alt = OUT.replace(".xlsx", "_NEW.xlsx")
        print(f"\n{OUT} is open in Excel - writing to {os.path.basename(alt)} instead.")
        OUT = alt
        main()
        print(f"\n!! {os.path.basename(alt)} is the CURRENT file. "
              f"Close Excel and rename it over the old one.")
