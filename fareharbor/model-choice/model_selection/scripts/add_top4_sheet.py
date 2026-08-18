"""
Add a Top_Models_Side_By_Side sheet to best_model_13.xlsx: the top 3 models by
coverage, plus any explicitly requested extras (ALSO_INCLUDE), plus the
incumbent gpt-4o-mini -- side by side per product per field.

The existing Side_By_Side sheet carries all 13 models, which is too wide to
read. This is the same data narrowed to the models a decision would actually
compare.

Column order is deliberate: raw source first (the source of truth), then the
candidates in coverage order, then gpt-4o-mini last -- so each candidate is
read against the raw text, and the incumbent sits at the end as the thing
being replaced.

The raw text is repeated on every row rather than being put on a separate
sheet, because judging placement means comparing a field's value against what
the supplier actually wrote, without switching sheets.

Rewrites the workbook via openpyxl (append, not regenerate) so the 6 existing
sheets keep their exact contents and styling.

Usage:
    python add_top4_sheet.py
"""
import sys
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from screen_model_comparison import ALL_FIELDS, PRODUCT_IDS

sys.stdout.reconfigure(encoding="utf-8")

TEST_DIR = Path(__file__).resolve().parent
XLSX = TEST_DIR / "best_model_13.xlsx"
SHEET = "Top_Models_Side_By_Side"
# the earlier 4-model name, removed on rewrite so the workbook never carries
# two near-identical sheets with a now-wrong count in the name
OLD_SHEETS = ["Top4_Side_By_Side"]

INCUMBENT = "gpt-4o-mini"
N_TOP = 3
# Explicitly requested alongside the top 3. Kept separate from N_TOP rather
# than widening it to 4, so "top N" stays a genuine ranking claim and this
# stays visibly a manual addition -- gpt-5.6-terra ranks 4th by coverage.
ALSO_INCLUDE = ["gpt-5.6-terra"]

CELL_LIMIT = 30000
ILLEGAL_XL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
BLUE_FILL = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
AMBER_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
RAW_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")


def clean(t):
    s = "" if t is None else str(t)
    s = ILLEGAL_XL_RE.sub("", s)
    return s[:CELL_LIMIT] if len(s) > CELL_LIMIT else s


def main():
    screen = json.loads((TEST_DIR / "bestmodel_screen_results.json").read_text(encoding="utf-8"))

    # rank by the same measure the Ranking sheet uses, so "top 3" matches it
    ranked = sorted(
        screen.keys(),
        key=lambda m: sum(v["word_coverage_pct"] for v in screen[m].values()) / len(screen[m]),
        reverse=True,
    )
    top = [m for m in ranked if m != INCUMBENT][:N_TOP]
    extra = [m for m in ALSO_INCLUDE if m not in top and m != INCUMBENT]
    missing = [m for m in extra if m not in screen]
    if missing:
        raise SystemExit(f"requested model(s) absent from screen results: {missing}")
    # candidates stay in coverage order; incumbent last as the thing replaced
    models = sorted(top + extra,
                    key=lambda m: ranked.index(m)) + [INCUMBENT]
    print(f"Top {N_TOP} by coverage: {', '.join(top)}")
    if extra:
        print(f"Also included (requested): {', '.join(extra)} "
              f"-- ranks {', '.join(str(ranked.index(m) + 1) for m in extra)} by coverage")
    print(f"Plus incumbent: {INCUMBENT}")

    cov = {m: round(sum(v["word_coverage_pct"] for v in screen[m].values())
                    / len(screen[m]), 2) for m in models}

    header = ["product_id", "field", "raw_description", "raw_booking_notes"] + models
    rows = []
    for pid in PRODUCT_IDS:
        first = screen[models[0]][pid]
        for field in ALL_FIELDS:
            vals = {m: str(screen[m][pid]["field_values"].get(field, "") or "").strip()
                    for m in models}
            # a field no model filled tells the reader nothing
            if not any(vals.values()):
                continue
            rows.append([pid, field, clean(first["raw_desc"]), clean(first["raw_booking"])]
                        + [clean(vals[m]) for m in models])

    wb = load_workbook(XLSX)
    for name in [SHEET] + OLD_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
    ws = wb.create_sheet(SHEET)

    ws.append(header)
    for r in rows:
        ws.append(r)

    n_cols = len(header)
    ws.freeze_panes = "C2"          # keep product_id + field visible while scrolling right
    ws.auto_filter.ref = ws.dimensions
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if 3 <= c <= 4:
            cell.fill = RAW_FILL
        elif header[c - 1] == INCUMBENT:
            cell.fill = BLUE_FILL
        else:
            cell.fill = HEADER_FILL

    widths = {1: 12, 2: 26, 3: 60, 4: 60}
    for c in range(5, n_cols + 1):
        widths[c] = 52
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Flag where the winner and the incumbent differ on a field.
    #
    # An empty cell on one side is NOT automatically a loss: the two models
    # often file the same sentence under different field names. Checked on this
    # data, all 11 cells gpt-4o-mini filled where gpt-5.4-nano was empty had
    # 100% word overlap with gpt-5.4-nano's output in some OTHER field -- they
    # are relocations, not dropped content. Colouring them red would have
    # claimed 11 losses that do not exist.
    #
    # So a cell is only marked red/green when the text is absent from the other
    # model's ENTIRE output for that product. Amber = present but filed
    # elsewhere, which is a placement difference for the judge, not a loss.
    win_i = header.index(top[0]) + 1
    inc_i = header.index(INCUMBENT) + 1

    def elsewhere(model, pid, text):
        """Is text present anywhere else in this model's output for pid?"""
        words = {w for w in re.sub(r"[^a-z0-9 ]", " ", text.lower()).split() if len(w) > 3}
        if not words:
            return True
        blob = " ".join(str(v or "") for v in screen[model][pid]["field_values"].values())
        have = {w for w in re.sub(r"[^a-z0-9 ]", " ", blob.lower()).split() if len(w) > 3}
        return len(words & have) / len(words) >= 0.9

    gained = lost = moved = 0
    for i in range(2, ws.max_row + 1):
        pid = str(ws.cell(row=i, column=1).value)
        w = str(ws.cell(row=i, column=win_i).value or "").strip()
        v = str(ws.cell(row=i, column=inc_i).value or "").strip()
        if w and not v:
            if elsewhere(INCUMBENT, pid, w):
                ws.cell(row=i, column=win_i).fill = AMBER_FILL
                moved += 1
            else:
                ws.cell(row=i, column=win_i).fill = GREEN_FILL
                gained += 1
        elif v and not w:
            if elsewhere(top[0], pid, v):
                ws.cell(row=i, column=inc_i).fill = AMBER_FILL
                moved += 1
            else:
                ws.cell(row=i, column=inc_i).fill = RED_FILL
                lost += 1

    # move the new sheet next to the existing wide one for discoverability
    order = wb.sheetnames
    order.remove(SHEET)
    at = order.index("Side_By_Side") + 1 if "Side_By_Side" in order else len(order)
    wb._sheets.insert(at, wb._sheets.pop(wb.sheetnames.index(SHEET)))

    # legend, appended below the data so the colours are readable without this
    # script's console output
    ws.append([])
    for text, fill in [
        (f"GREEN = {top[0]} captured content {INCUMBENT} lost entirely", GREEN_FILL),
        (f"RED = {INCUMBENT} captured content {top[0]} lost entirely", RED_FILL),
        ("AMBER = same content, filed under a different field name "
         "(placement difference, NOT lost content)", AMBER_FILL),
        ("Coverage counts words that survived ANYWHERE; it cannot say whether a "
         "field is the right one. Compare against the raw columns to judge placement.",
         None),
    ]:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=True)
        if fill:
            c.fill = fill

    wb.save(XLSX)

    print(f"\nWrote sheet '{SHEET}' -> {XLSX.name}")
    print(f"  {len(rows)} rows ({len(PRODUCT_IDS)} products x fields any model filled)")
    print(f"  columns: raw_description, raw_booking_notes, {', '.join(models)}")
    print(f"\nCoverage of the {len(models)} shown:")
    for m in models:
        tag = "  <- current" if m == INCUMBENT else ""
        print(f"  {m:<16}{cov[m]:>7}%{tag}")
    print(f"\nField-level diff, {top[0]} vs {INCUMBENT}:")
    print(f"  GREEN {gained:>3}  content {top[0]} captured that {INCUMBENT} lost entirely")
    print(f"  RED   {lost:>3}  content {INCUMBENT} captured that {top[0]} lost entirely")
    print(f"  AMBER {moved:>3}  same content, different field -- placement difference, "
          f"not a loss")


if __name__ == "__main__":
    main()
