"""
Build best_model_13.xlsx and best_model_13_review.txt.

The decision-relevant sheet is Quality_vs_Cost: coverage plotted against cost,
so the point where quality stops improving is visible directly.

Judge columns are absent by design -- no judge has been run on these models.

Usage:
    python build_best_model_workbook.py
"""
import sys
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from screen_model_comparison import ALL_FIELDS, PRODUCT_IDS
from build_best_model_batches import CANDIDATES, EXISTING_BASELINES

sys.stdout.reconfigure(encoding="utf-8")

TEST_DIR = Path(__file__).resolve().parent
OUT_XLSX = TEST_DIR / "best_model_13.xlsx"
OUT_TXT = TEST_DIR / "best_model_13_review.txt"

CELL_LIMIT = 30000
ILLEGAL_XL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
MD_RE = re.compile(r"\*\*|^#{1,3}\s", re.MULTILINE)

FAREHARBOR_N = 11236
ALL_SOURCES_N = 23034
# measured token shape per product from the gpt-5.5-pro 50-product run
PT_PER_PRODUCT, CT_PER_PRODUCT = 6705, 2408

# official Batch API rates, USD per 1M tokens (input, output)
PRICING = {
    "gpt-5-nano": (0.025, 0.20), "gpt-4.1-nano": (0.05, 0.20),
    "gpt-4o-mini": (0.075, 0.30), "gpt-5.6-luna": (0.10, 0.60),
    "gpt-5.4-nano": (0.10, 0.625), "gpt-5-mini": (0.125, 1.00),
    "gpt-4.1-mini": (0.20, 0.80), "gpt-5.4-mini": (0.375, 2.25),
    "o4-mini": (0.55, 2.20), "gpt-5": (0.625, 5.00),
    "gpt-5.6-terra": (1.00, 6.00), "gpt-5.4": (1.25, 7.50),
    "gpt-5.5-pro": (15.00, 90.00),
}

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

RULE = "=" * 100
THIN = "-" * 100


def clean(t):
    s = "" if t is None else str(t)
    s = ILLEGAL_XL_RE.sub("", s)
    return s[:CELL_LIMIT] if len(s) > CELL_LIMIT else s


def indent(text, prefix="    "):
    return prefix + "(empty)" if not str(text).strip() else "\n".join(
        prefix + ln for ln in str(text).split("\n"))


def cost_for(model, n_products):
    if model not in PRICING:
        return None
    pin, pout = PRICING[model]
    return (PT_PER_PRODUCT / 1e6 * pin + CT_PER_PRODUCT / 1e6 * pout) * n_products


def style(ws, n_cols, widths=None, wrap=None):
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = Font(bold=True)
    for c, w in (widths or {}).items():
        if c <= n_cols:
            ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top") \
                if (wrap and cell.column in wrap) else Alignment(vertical="top")


def main():
    screen = json.loads((TEST_DIR / "bestmodel_screen_results.json").read_text(encoding="utf-8"))
    try:
        batch = json.loads((TEST_DIR / "bestmodel_batch_results.json").read_text(encoding="utf-8"))
    except FileNotFoundError:
        batch = {}

    models = [m for m in CANDIDATES + EXISTING_BASELINES if m in screen]
    baseline_cov = None

    # ---------- Ranking ----------
    rank_rows = []
    for model in models:
        per = screen[model]
        vals = [per[p] for p in PRODUCT_IDS if p in per]
        cov = round(sum(v["word_coverage_pct"] for v in vals) / len(vals), 2)
        if model == "gpt-4o-mini":
            baseline_cov = cov
        empty_fields = round(sum(
            sum(1 for x in v["field_values"].values() if not str(x).strip())
            for v in vals) / len(vals), 1)
        md_fields = sum(1 for v in vals for x in v["field_values"].values()
                        if x and MD_RE.search(str(x)))
        b = batch.get(model, {})
        rank_rows.append({
            "model": model,
            "tested": "NEW" if model in CANDIDATES else "existing baseline",
            "avg_coverage_pct": cov,
            "total_MISSING": sum(v["units_missing"] for v in vals),
            "total_PARTIAL": sum(v["units_partial"] for v in vals),
            "avg_fields_filled": round(sum(
                sum(1 for x in v["field_values"].values() if str(x).strip())
                for v in vals) / len(vals), 1),
            "avg_fields_EMPTY": empty_fields,
            "markdown_junk_fields": md_fields,
            "truncated": sum(1 for v in vals if v.get("truncated")),
            "bad_json": sum(len(v.get("bad_json", [])) for v in vals),
            "cost_11236_usd": (lambda c: round(c, 2) if c is not None else "n/a")(
                cost_for(model, FAREHARBOR_N)),
            "cost_23034_usd": (lambda c: round(c, 2) if c is not None else "n/a")(
                cost_for(model, ALL_SOURCES_N)),
            "wall_clock_s": b.get("wall_clock_seconds", ""),
            "prompt_tokens": b.get("prompt_tokens", ""),
            "completion_tokens": b.get("completion_tokens", ""),
        })
    rank_df = pd.DataFrame(rank_rows).sort_values("avg_coverage_pct", ascending=False)

    # ---------- Quality vs Cost ----------
    qc = rank_df[["model", "avg_coverage_pct", "total_MISSING",
                  "cost_11236_usd", "cost_23034_usd"]].copy()
    qc["vs_gpt4omini_coverage"] = qc["avg_coverage_pct"].map(
        lambda c: round(c - baseline_cov, 2) if baseline_cov else "")
    base_cost = cost_for("gpt-4o-mini", ALL_SOURCES_N)
    qc["cost_multiple_vs_4omini"] = qc["model"].map(
        lambda m: round(cost_for(m, ALL_SOURCES_N) / base_cost, 1)
        if cost_for(m, ALL_SOURCES_N) else "n/a")
    qc = qc.sort_values("cost_23034_usd")

    # ---------- Per_Product ----------
    pp_rows = []
    for pid in PRODUCT_IDS:
        row = {"product_id": pid,
               "input_words": screen[models[0]][pid]["input_words"]}
        for model in models:
            d = screen[model].get(pid, {})
            row[f"{model}_cov"] = d.get("word_coverage_pct")
            row[f"{model}_miss"] = d.get("units_missing")
        pp_rows.append(row)
    pp_df = pd.DataFrame(pp_rows)

    # ---------- Side_By_Side / Raw / Content ----------
    sbs_rows, raw_rows, content_rows = [], [], []
    for pid in PRODUCT_IDS:
        first = screen[models[0]][pid]
        raw_rows.append({"product_id": pid, "input_words": first["input_words"],
                         "raw_description": clean(first["raw_desc"]),
                         "raw_booking_notes": clean(first["raw_booking"])})
        for field in ALL_FIELDS:
            vals = {m: str(screen[m][pid]["field_values"].get(field, "") or "").strip()
                    for m in models}
            if not any(vals.values()):
                continue
            row = {"product_id": pid, "field": field,
                   "n_models_filled": sum(1 for v in vals.values() if v)}
            row.update({m: clean(v) for m, v in vals.items()})
            sbs_rows.append(row)
        for model in models:
            content_rows.append({"product_id": pid, "model": model,
                                 **{f: clean(screen[model][pid]["field_values"].get(f, ""))
                                    for f in ALL_FIELDS}})

    sbs_df = pd.DataFrame(sbs_rows)
    raw_df = pd.DataFrame(raw_rows)
    content_df = pd.DataFrame(content_rows)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        rank_df.to_excel(w, index=False, sheet_name="Ranking")
        qc.to_excel(w, index=False, sheet_name="Quality_vs_Cost")
        pp_df.to_excel(w, index=False, sheet_name="Per_Product")
        sbs_df.to_excel(w, index=False, sheet_name="Side_By_Side")
        raw_df.to_excel(w, index=False, sheet_name="Raw_Source")
        content_df.to_excel(w, index=False, sheet_name="Content_By_Model")

        ws = w.sheets["Ranking"]
        style(ws, len(rank_df.columns),
              {1: 16, 2: 18, 3: 17, 4: 14, 5: 14, 6: 17, 7: 17, 8: 20,
               9: 11, 10: 10, 11: 16, 12: 16, 13: 13, 14: 14, 15: 17})
        ci = rank_df.columns.get_loc("avg_coverage_pct") + 1
        mi = rank_df.columns.get_loc("model") + 1
        for i, r in enumerate(rank_df.itertuples(), start=2):
            v = getattr(r, "avg_coverage_pct")
            ws.cell(row=i, column=ci).fill = (
                GREEN_FILL if v >= 98 else AMBER_FILL if v >= 93 else RED_FILL)
            if getattr(r, "model") == "gpt-4o-mini":
                ws.cell(row=i, column=mi).fill = BLUE_FILL

        ws = w.sheets["Quality_vs_Cost"]
        style(ws, len(qc.columns), {1: 16, 2: 17, 3: 14, 4: 16, 5: 16, 6: 22, 7: 24})
        for i, r in enumerate(qc.itertuples(), start=2):
            d = getattr(r, "vs_gpt4omini_coverage")
            if isinstance(d, (int, float)):
                ws.cell(row=i, column=6).fill = (
                    GREEN_FILL if d > 0 else BLUE_FILL if d == 0 else RED_FILL)

        style(w.sheets["Per_Product"], len(pp_df.columns), {1: 12, 2: 12})
        style(w.sheets["Side_By_Side"], len(sbs_df.columns),
              {1: 12, 2: 28, 3: 15, **{i: 46 for i in range(4, len(sbs_df.columns) + 1)}},
              set(range(4, len(sbs_df.columns) + 1)))
        style(w.sheets["Raw_Source"], len(raw_df.columns),
              {1: 12, 2: 13, 3: 110, 4: 110}, {3, 4})
        style(w.sheets["Content_By_Model"], len(content_df.columns),
              {1: 12, 2: 16, **{i: 48 for i in range(3, len(content_df.columns) + 1)}},
              set(range(3, len(content_df.columns) + 1)))

    # ---------- review txt ----------
    lines = [RULE, f"BEST-MODEL COMPARISON -- {len(models)} models, "
             f"{len(PRODUCT_IDS)} identical products", RULE, "",
             "Raw supplier text appears above each product, then every model's fields in",
             "turn. No paste fix, no judge -- this is raw extraction output.", "",
             "Ranked by average word coverage:"]
    for r in rank_df.itertuples():
        lines.append(f"   {getattr(r, 'model'):<16} {getattr(r, 'avg_coverage_pct'):>6}%   "
                     f"MISSING {getattr(r, 'total_MISSING'):<4} "
                     f"${getattr(r, 'cost_23034_usd')} for 23,034 products")

    for pid in PRODUCT_IDS:
        first = screen[models[0]][pid]
        lines += ["", RULE, f"PRODUCT {pid}   ({first['input_words']} words)", RULE,
                  "", THIN, "RAW DESCRIPTION (source of truth)", THIN,
                  indent(first["raw_desc"]),
                  "", THIN, "RAW BOOKING NOTES (source of truth)", THIN,
                  indent(first["raw_booking"])]
        for model in models:
            d = screen[model][pid]
            lines += ["", THIN,
                      f"MODEL: {model}   coverage {d['word_coverage_pct']}%   "
                      f"missing {d['units_missing']}   partial {d['units_partial']}",
                      THIN]
            for f in ALL_FIELDS:
                v = str(d["field_values"].get(f, "") or "").strip()
                if v:
                    lines += [f"  [{f}]", indent(v, "      ")]
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    print(f"Wrote {OUT_XLSX.name}")
    print(f"  Ranking({len(rank_df)}) Quality_vs_Cost({len(qc)}) Per_Product({len(pp_df)}) "
          f"Side_By_Side({len(sbs_df)}) Raw_Source({len(raw_df)}) "
          f"Content_By_Model({len(content_df)})")
    print(f"Wrote {OUT_TXT.name} ({len(lines)} lines)")
    print()
    print(rank_df[["model", "tested", "avg_coverage_pct", "total_MISSING",
                   "avg_fields_EMPTY", "cost_23034_usd"]].to_string(index=False))
    return rank_df


if __name__ == "__main__":
    main()
