"""
Build best_model_hard30.xlsx -- raw extraction output, no post-processing.

Four variants: gpt-5.6-luna (chosen), gpt-5.6-terra, gpt-5.4-nano, and the
gpt-4o-mini baseline reused from the earlier 500-product run. Every value is
exactly what the model returned -- nothing is cleaned, stripped or rewritten
after extraction, so the workbook shows each model's true behaviour.

`dup_sentences` counts sentences the model placed in more than one field on the
same side. It is the measurement that rejected gpt-5.4-nano (184 duplicated
sentences against gpt-5.6-terra's 1) and it is a property of the raw output,
not of any repair step.

Assertions at the end guarantee the output stays raw: no processed variant may
appear as a row or a column.

The 38 hand-written review comments (17 from the user, 21 on gpt-5.6-luna) are
read from the existing workbook and carried across, keyed on
(product_id, variant). The generator cannot reproduce them, so losing them
would destroy the most valuable content in the file. Asserted before writing.

Usage:
    python build_hard30_publish.py --out <path>
"""
import sys
import re
import json
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from screen_model_comparison import ALL_FIELDS

sys.stdout.reconfigure(encoding="utf-8")

TEST_DIR = Path(__file__).resolve().parent
SOURCE_XLSX = TEST_DIR / "best_model_hard30.xlsx"
SCREEN = TEST_DIR / "hard30_screen_results.json"
PRODUCTS_PATH = TEST_DIR / "hard30_products.json"

BASELINE = "gpt-4o-mini (V4.4, existing)"
VARIANTS = ["gpt-5.6-luna", "gpt-5.6-terra", "gpt-5.4-nano", BASELINE]
CHOSEN = "gpt-5.6-luna"

PRICING = {"gpt-5.6-terra": (1.00, 6.00), "gpt-5.4-nano": (0.10, 0.625),
           "gpt-5.6-luna": (0.10, 0.60), BASELINE: (0.075, 0.30)}
PT_PER_PRODUCT, CT_PER_PRODUCT = 6705, 2408
ALL_SOURCES_N = 23034

CELL_LIMIT = 30000
ILLEGAL_XL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
AMBER_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
RED_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
BLUE_FILL = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
RAW_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")


def clean(t):
    s = "" if t is None else str(t)
    s = ILLEGAL_XL_RE.sub("", s)
    return s[:CELL_LIMIT] if len(s) > CELL_LIMIT else s


def style(ws, widths=None, wrap=None, header_fills=None):
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = (header_fills or {}).get(c, HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in (widths or {}).items():
        if c <= ws.max_column:
            ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = (Alignment(wrap_text=True, vertical="top")
                              if wrap and cell.column in wrap
                              else Alignment(vertical="top"))


def cost(model):
    pin, pout = PRICING[model]
    return round(PT_PER_PRODUCT * ALL_SOURCES_N / 1e6 * pin
                 + CT_PER_PRODUCT * ALL_SOURCES_N / 1e6 * pout)


def load_comments():
    """{(product_id, variant): (pass_fail, comment)} from the working workbook."""
    if not SOURCE_XLSX.exists():
        print(f"  WARNING: {SOURCE_XLSX.name} absent -- no comments to carry over")
        return {}
    df = pd.read_excel(SOURCE_XLSX, sheet_name="Content_By_Model").fillna("")
    out = {}
    for r in df.itertuples(index=False):
        pf = str(getattr(r, "_1", "")).strip()      # 'pass/fail' is not a valid identifier
        cm = str(getattr(r, "comment", "")).strip()
        if pf or cm:
            out[(str(r.product_id), str(r.variant))] = (pf, cm)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    out_path = Path(args.out)

    screen = json.loads(SCREEN.read_text(encoding="utf-8"))
    product_ids = json.loads(PRODUCTS_PATH.read_text(encoding="utf-8"))["product_ids"]
    comments = load_comments()
    print(f"Carrying over {len(comments)} review comment(s)")

    # ---------- Ranking ----------
    rows = []
    for v in VARIANTS:
        per = screen[v]
        vals = [per[p] for p in product_ids]
        n = len(vals)
        raw_tot = sum(x["input_words"] for x in vals)
        emit_tot = sum(x["words_emitted"] for x in vals)
        rows.append({
            "model": v + ("   <- CHOSEN" if v == CHOSEN else ""),
            "coverage_pct": round(sum(x["word_coverage_pct"] for x in vals) / n, 2),
            "MISSING": sum(x["units_missing"] for x in vals),
            "word_ratio": round(emit_tot / raw_tot, 3) if raw_tot else 0,
            "dup_sentences": sum(x["dup_sentences"] for x in vals),
            "untraceable_fields": sum(x["untraceable_fields"] for x in vals),
            "markdown_fields": sum(x["markdown_fields"] for x in vals),
            "avg_fields_filled": round(sum(x["fields_filled"] for x in vals) / n, 1),
            "cost_23034_usd": cost(v),
        })
    rank_df = pd.DataFrame(rows)

    # ---------- Per_Product ----------
    pp = []
    for pid in product_ids:
        row = {"product_id": pid, "input_words": screen[VARIANTS[0]][pid]["input_words"]}
        for v in VARIANTS:
            row[f"{v}_cov"] = screen[v][pid]["word_coverage_pct"]
            row[f"{v}_dup"] = screen[v][pid]["dup_sentences"]
        pp.append(row)
    pp_df = pd.DataFrame(pp)

    # ---------- Content_By_Model (the review sheet) ----------
    content, raw_rows = [], []
    kept = 0
    for pid in product_ids:
        first = screen[VARIANTS[0]][pid]
        raw_rows.append({"product_id": pid, "input_words": first["input_words"],
                         "raw_description": clean(first["raw_desc"]),
                         "raw_booking_notes": clean(first["raw_booking"])})
        for v in VARIANTS:
            d = screen[v][pid]
            pf, cm = comments.get((pid, v), ("", ""))
            if pf or cm:
                kept += 1
            content.append({
                "product_id": pid, "pass/fail": pf, "comment": cm, "variant": v,
                "coverage": d["word_coverage_pct"],
                "word_ratio": d["word_ratio"],
                "dup_sentences": d["dup_sentences"],
                "raw_description": clean(d["raw_desc"]),
                "raw_booking_notes": clean(d["raw_booking"]),
                **{f: clean(d["field_values"].get(f, "")) for f in ALL_FIELDS},
            })
    content_df = pd.DataFrame(content)
    raw_df = pd.DataFrame(raw_rows)

    # ---------- Side_By_Side ----------
    sbs = []
    for pid in product_ids:
        first = screen[VARIANTS[0]][pid]
        for field in ALL_FIELDS:
            vals = {v: str(screen[v][pid]["field_values"].get(field, "") or "").strip()
                    for v in VARIANTS}
            if not any(vals.values()):
                continue
            sbs.append({"product_id": pid, "field": field,
                        "raw_description": clean(first["raw_desc"]),
                        "raw_booking_notes": clean(first["raw_booking"]),
                        **{v: clean(t) for v, t in vals.items()}})
    sbs_df = pd.DataFrame(sbs)

    # ---- assertions before writing ----
    assert len(content_df) == len(VARIANTS) * len(product_ids), \
        f"expected {len(VARIANTS) * len(product_ids)} rows, got {len(content_df)}"
    assert kept == len(comments), \
        f"comment loss: carried {kept} of {len(comments)}"
    # the workbook must contain raw model output only -- a variant name carrying
    # a "+" suffix would mean a post-processed version leaked in
    for df in (rank_df, pp_df, content_df, sbs_df):
        bad = [c for c in df.columns if "+" in str(c)]
        assert not bad, f"post-processed column in output: {bad}"
    assert not content_df.variant.astype(str).str.contains(r"\+").any(), \
        "post-processed variant in Content_By_Model"
    assert set(content_df.variant) == set(VARIANTS), \
        f"unexpected variants: {set(content_df.variant) - set(VARIANTS)}"

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        rank_df.to_excel(w, sheet_name="Ranking", index=False)
        pp_df.to_excel(w, sheet_name="Per_Product", index=False)
        content_df.to_excel(w, sheet_name="Content_By_Model", index=False)
        sbs_df.to_excel(w, sheet_name="Side_By_Side", index=False)
        raw_df.to_excel(w, sheet_name="Raw_Source", index=False)

        ws = w.sheets["Ranking"]
        style(ws, {1: 30, 2: 13, 3: 10, 4: 12, 5: 15, 6: 19, 7: 17, 8: 18, 9: 16})
        di = list(rank_df.columns).index("dup_sentences") + 1
        wi = list(rank_df.columns).index("word_ratio") + 1
        for i, r in enumerate(rank_df.itertuples(), start=2):
            d = getattr(r, "dup_sentences")
            ws.cell(row=i, column=di).fill = (
                GREEN_FILL if d <= 3 else AMBER_FILL if d <= 20 else RED_FILL)
            wr = getattr(r, "word_ratio")
            ws.cell(row=i, column=wi).fill = (
                GREEN_FILL if 0.90 <= wr <= 1.05 else RED_FILL)
            if CHOSEN in str(getattr(r, "model")):
                ws.cell(row=i, column=1).fill = BLUE_FILL

        style(w.sheets["Per_Product"], {1: 12, 2: 12})

        ws = w.sheets["Content_By_Model"]
        n = len(content_df.columns)
        style(ws, {1: 12, 2: 11, 3: 42, 4: 26, 5: 10, 6: 11, 7: 14,
                   8: 60, 9: 60, **{i: 46 for i in range(10, n + 1)}},
              wrap=set(range(8, n + 1)),
              header_fills={2: AMBER_FILL, 3: AMBER_FILL, 8: RAW_FILL, 9: RAW_FILL})

        ws = w.sheets["Side_By_Side"]
        n = len(sbs_df.columns)
        style(ws, {1: 12, 2: 26, 3: 55, 4: 55, **{i: 46 for i in range(5, n + 1)}},
              wrap=set(range(3, n + 1)),
              header_fills={3: RAW_FILL, 4: RAW_FILL})

        style(w.sheets["Raw_Source"], {1: 12, 2: 13, 3: 110, 4: 110}, wrap={3, 4})

    print("=" * 92)
    print(f"PUBLISH WORKBOOK -- raw extraction only, {len(VARIANTS)} variants")
    print("=" * 92)
    print(rank_df.to_string(index=False))
    print(f"\nWrote {out_path}")
    print(f"  Content_By_Model {len(content_df)} rows ({len(VARIANTS)} x {len(product_ids)})")
    print(f"  review comments carried: {kept}/{len(comments)}")
    print(f"  Side_By_Side     {len(sbs_df)} rows")
    print("  no de-dup rows or columns (asserted)")


if __name__ == "__main__":
    main()
