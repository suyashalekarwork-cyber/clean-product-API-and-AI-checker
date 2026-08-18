"""Step 1 vs Step 2, stacked, for the same 100 Rezdy products.

Every field gets TWO rows: what Step 1 produced, and directly beneath it what
Step 2 produced. Same product, same supplier text, so any difference is the
method and nothing else.

THE ONE THING THIS WORKBOOK MUST NOT DO IS FLATTER STEP 2, so three honesty
rules are built into the layout:

  1. STEP 2 FILLS FEWER FIELDS, AND THAT IS THE POINT. Step 1 filled highlights
     on 97 of 100 products; Step 2 fills 45. Step 1 was guessing -- it put text
     under a heading the supplier never wrote, which is how 294 of 329 sampled
     failures happened. A blank in Step 2 means "the supplier wrote no heading",
     and the text is in `about`, not lost. The workbook says so on every blank.

  2. FIELDS STEP 1 NEVER HAD ARE NOT A STEP 2 WIN. Step 1 shipped 9 fields;
     Step 2 ships 22. For the other 13 there is nothing to compare against, and
     they are labelled "not in Step 1" rather than scored as an improvement.

  3. FIELDS STEP 2 DROPPED ARE SHOWN TOO. Step 1's `desc_other` and
     `desc_booking_notes` have no Step 2 equivalent by design -- V5 has no
     "other" field. Their content is not missing; it moved to `about`. Shown so
     a reviewer can confirm that rather than take it on trust.

openpyxl raises IllegalCharacterError on the ASCII control characters present in
some supplier text, and Excel truncates a cell over 32,767 characters -- both are
handled in clean() rather than discovered during a demo.
"""
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
T = ROOT / "data_pipeline" / "batch_api_test"
sys.path.insert(0, str(T))

from booking_common import parse_booking_json          # noqa: E402
from rezdy_common import RAW_DIR, html_to_markdown     # noqa: E402

OUT = ROOT / "exports" / "rezdy_step1_vs_step2_100.xlsx"
STEP1_CSV = ROOT / "data_pipeline" / "stage1" / "rezdy_master.csv"
STEP2_JSONL = T / "rezdy_desc_100_output.jsonl"
PRODUCTS = T / "rezdy_desc_100_products.json"

# (step1 column, step2 column, label). None = the side has no such field.
FIELDS = [
    ("desc_about", "redo_desc_about", "About / description"),
    ("desc_highlights", "redo_desc_highlights", "Highlights"),
    ("desc_what_included", "redo_desc_what_included", "What's included"),
    ("desc_what_excluded", "redo_desc_what_excluded", "What's not included"),
    ("desc_itinerary", "redo_desc_itinerary", "Itinerary"),
    ("desc_what_to_bring", "redo_desc_what_to_bring", "What to bring"),
    ("desc_cancellation", "redo_desc_cancellation", "Cancellation"),
    ("desc_booking_notes", None, "Booking notes"),
    ("desc_other", None, "Other"),
    (None, "redo_desc_important_info", "Important info"),
    (None, "redo_meeting_point", "Meeting point"),
    (None, "redo_desc_restrictions", "Restrictions"),
    (None, "redo_desc_duration_text", "Duration"),
    (None, "redo_desc_extras", "Extras"),
    (None, "redo_desc_pricing", "Pricing"),
    (None, "redo_desc_check_in", "Check in"),
    (None, "redo_desc_health_safety", "Health & safety"),
    (None, "redo_desc_contact", "Contact"),
    (None, "redo_desc_faqs", "FAQs"),
    (None, "redo_desc_disclaimers", "Disclaimers"),
    (None, "redo_desc_accessibility", "Accessibility"),
    (None, "redo_group_size", "Group size"),
    (None, "redo_desc_special_requirements", "Special requirements"),
]

CTRL = re.compile(r"[\000-\010\013\014\016-\037]")
MAXCELL = 32000

HDR = PatternFill("solid", fgColor="1F3864")
S1_FILL = PatternFill("solid", fgColor="FCE4D6")      # amber  = Step 1
S2_FILL = PatternFill("solid", fgColor="DDEBF7")      # blue   = Step 2
PROD_FILL = PatternFill("solid", fgColor="2F5597")
NA_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean(v):
    s = "" if v is None else str(v)
    if s.strip().lower() == "nan":
        s = ""
    s = CTRL.sub("", s)
    return s[:MAXCELL]


def filled(v):
    return bool(clean(v).strip())


def load():
    step2 = {}
    for line in STEP2_JSONL.open(encoding="utf-8"):
        r = json.loads(line)
        f, _ = parse_booking_json(
            r["response"]["body"]["choices"][0]["message"]["content"])
        step2[r["custom_id"].split("|")[0]] = f or {}

    d = pd.read_csv(STEP1_CSV, low_memory=False)
    d["product_id"] = d.product_id.astype(str)
    step1 = d[d.product_id.isin(step2)].set_index("product_id")

    meta = {p["product_id"]: p
            for p in json.loads(PRODUCTS.read_text(encoding="utf-8"))}
    return step1, step2, meta


def raw_of(pid):
    hits = list(RAW_DIR.glob(f"Rezdy-*-{pid}.json"))
    if not hits:
        return ""
    p = json.loads(hits[0].read_text(encoding="utf-8")).get("product", {})
    return html_to_markdown(p.get("description") or "")


def verdict(c1, c2, v1, v2):
    if c1 is None:
        return ("not in Step 1" if filled(v2)
                else "not in Step 1 / no heading")
    if c2 is None:
        return "no Step 2 field (moved to About)"
    if filled(v1) and filled(v2):
        return "both"
    if filled(v1) and not filled(v2):
        return "STEP 1 only -- no supplier heading"
    if filled(v2) and not filled(v1):
        return "STEP 2 only"
    return "both empty"


def summary_sheet(wb, step1, step2, meta):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Rezdy: Step 1 vs Step 2"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws["A2"] = ("The same 100 products, the same supplier text, two methods. "
                "Step 1 sorted text by what it looked like. Step 2 only fills a "
                "field when the supplier wrote a heading for it.")
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:F2")

    # ---- the headline: how much of the supplier's words each method kept
    W = re.compile(r"[A-Za-z0-9']+")
    raw_n = s1_n = s2_n = 0
    for pid in step2:
        R = set(W.findall(raw_of(pid).lower()))
        a = " ".join(clean(step1.loc[pid, c]) for c, _, _ in FIELDS
                     if c and c in step1.columns)
        b = " ".join(clean(v) for k, v in step2[pid].items() if k != "redo_flags")
        raw_n += len(R)
        s1_n += len(R & set(W.findall(a.lower())))
        s2_n += len(R & set(W.findall(b.lower())))

    r = 4
    ws.cell(r, 1, "How much of the supplier's text survived").font = Font(
        bold=True, size=12)
    r += 1
    for label, val, note in [
        ("Step 1", f"{100*s1_n/raw_n:.1f}%", "more than a quarter of the "
         "supplier's words reached no field at all"),
        ("Step 2", f"{100*s2_n/raw_n:.1f}%", "almost nothing is dropped"),
    ]:
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, val)
        c.font = Font(bold=True, size=14,
                      color="C00000" if label == "Step 1" else "1F7A1F")
        ws.cell(r, 3, note).font = Font(italic=True, color="595959")
        r += 1

    # ---- per-field fill counts
    r += 1
    ws.cell(r, 1, "Fields filled, out of 100 products").font = Font(
        bold=True, size=12)
    r += 1
    heads = ["Field", "Step 1", "Step 2", "How to read it"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(r, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR
        c.border = BOX
    r += 1
    for c1, c2, label in FIELDS:
        n1 = (sum(filled(step1.loc[p, c1]) for p in step2 if p in step1.index)
              if c1 and c1 in step1.columns else None)
        n2 = (sum(filled(step2[p].get(c2)) for p in step2) if c2 else None)
        if c1 is None:
            note = "Step 1 had no such field -- not an improvement, just new"
        elif c2 is None:
            note = "Step 2 has no such field by design; content moved to About"
        elif n2 < n1:
            note = (f"Step 2 fills {n1-n2} fewer -- those products had no "
                    f"heading, so the text stayed in About")
        else:
            note = "Step 2 fills the same or more"
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, "--" if n1 is None else n1).alignment = Alignment(
            horizontal="center")
        ws.cell(r, 3, "--" if n2 is None else n2).alignment = Alignment(
            horizontal="center")
        ws.cell(r, 4, note)
        if n1 is None:
            ws.cell(r, 2).fill = NA_FILL
        if n2 is None:
            ws.cell(r, 3).fill = NA_FILL
        for i in range(1, 5):
            ws.cell(r, i).border = BOX
        r += 1

    r += 1
    ws.cell(r, 1, "Reading a blank in Step 2").font = Font(bold=True, size=12)
    r += 1
    for line in [
        "A blank does NOT mean the text was lost.",
        "It means the supplier never wrote a heading for that field, so the "
        "text stayed in About -- which is the correct answer.",
        "Step 1 filled those fields by guessing, and that guessing is what put "
        "294 of 329 sampled failures under the wrong heading.",
    ]:
        ws.cell(r, 1, line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    for col, w in zip("ABCDEF", [26, 10, 10, 70, 14, 14]):
        ws.column_dimensions[col].width = w
    return ws


def detail_sheet(wb, step1, step2, meta):
    ws = wb.create_sheet("Step 1 vs Step 2")
    ws.sheet_view.showGridLines = False
    heads = ["Field", "Method", "What it produced", "Words", "Verdict"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR
        c.border = BOX
    ws.freeze_panes = "A2"

    r = 2
    for pid in sorted(step2, key=lambda p: -meta.get(p, {}).get("n_headings", 0)):
        m = meta.get(pid, {})
        c = ws.cell(r, 1, f"{pid}   {m.get('supplier','')}   "
                          f"{m.get('name','')[:70]}   "
                          f"[{m.get('n_headings',0)} headings, "
                          f"{m.get('words',0):,} words]")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PROD_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        for i in range(1, 6):
            ws.cell(r, i).fill = PROD_FILL
        r += 1

        for c1, c2, label in FIELDS:
            v1 = clean(step1.loc[pid, c1]) if (c1 and c1 in step1.columns
                                               and pid in step1.index) else ""
            v2 = clean(step2[pid].get(c2)) if c2 else ""
            if not (filled(v1) or filled(v2)):
                continue                       # both empty -- nothing to show
            v = verdict(c1, c2, v1, v2)

            for method, val, fill, exists in (
                    ("STEP 1", v1, S1_FILL, c1 is not None),
                    ("STEP 2", v2, S2_FILL, c2 is not None)):
                ws.cell(r, 1, label if method == "STEP 1" else "")
                ws.cell(r, 2, method).font = Font(bold=True, size=9)
                if not exists:
                    ws.cell(r, 3, "— this method has no such field —").font = \
                        Font(italic=True, color="808080")
                elif not filled(val):
                    # "no supplier heading" explains a STEP 2 blank only. Step 1
                    # never looked at headings, so using that wording for its
                    # blanks would explain one method's behaviour with the
                    # other's logic -- and flatter Step 2 by implying its rule
                    # was already being followed.
                    ws.cell(r, 3, "(empty — supplier wrote no heading for this)"
                            if method == "STEP 2" else "(empty)").font = \
                        Font(italic=True, color="808080")
                else:
                    ws.cell(r, 3, val)
                ws.cell(r, 4, len(val.split()) if filled(val) else 0)
                ws.cell(r, 5, v if method == "STEP 1" else "")
                for i in range(1, 6):
                    cell = ws.cell(r, i)
                    cell.fill = NA_FILL if not exists else fill
                    cell.border = BOX
                    cell.alignment = Alignment(vertical="top", wrap_text=(i == 3))
                r += 1

        flags = clean(step2[pid].get("redo_flags"))
        if flags:
            ws.cell(r, 1, "Flags").font = Font(bold=True)
            ws.cell(r, 2, "STEP 2").font = Font(bold=True, size=9)
            ws.cell(r, 3, flags).alignment = Alignment(vertical="top",
                                                       wrap_text=True)
            ws.cell(r, 5, "model's own notes")
            for i in range(1, 6):
                ws.cell(r, i).fill = PatternFill("solid", fgColor="FFF2CC")
                ws.cell(r, i).border = BOX
            r += 1
        r += 1                                  # blank line between products

    for col, w in zip("ABCDE", [22, 9, 110, 8, 34]):
        ws.column_dimensions[col].width = w
    return ws


def main():
    step1, step2, meta = load()
    print(f"step 1 rows: {len(step1)}   step 2 products: {len(step2)}")

    wb = Workbook()
    wb.remove(wb.active)
    summary_sheet(wb, step1, step2, meta)
    ws = detail_sheet(wb, step1, step2, meta)
    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)

    print(f"sheets      : {wb.sheetnames}")
    print(f"detail rows : {ws.max_row:,}")
    print(f"\nwrote {OUT}")


if __name__ == "__main__":
    main()
