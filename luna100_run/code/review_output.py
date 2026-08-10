"""
Review an extraction run against the RAW supplier text.

The raw text is the only reference. Nothing here compares one prompt version to
another -- a version can look good purely because the previous one was worse, and
the previous one was often measurably wrong.

    python review_output.py                                  # newest run in output/
    python review_output.py --run luna100_v4_8_2_output.jsonl

Produces:
    reports/section_review.xlsx   Read_Me / Summary / Checks / Itinerary / FAQ /
                                  Whats_Included / All_Products
    reports/SECTION_STATUS.md     scorecard, one row per known issue

Sheets carry YOUR VERDICT / YOUR COMMENT columns, blank, with the raw text beside
the extracted value so a verdict can be checked without leaving the row.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "output"
REPORTS = ROOT / "reports"

MODEL = "gpt-5.6-luna"
ITIN, FAQ, INC = ("redo_desc_itinerary", "redo_desc_faqs",
                  "redo_desc_what_included")
REPAIRED = []

HDR = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
RAW = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
REV = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
RED = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
CELL_LIMIT = 32000
ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

CLOCK_START = re.compile(r"^\s*\w{0,14}:?\s*\d{1,2}\s?[:h.]\s?\d{2}\s?(am|pm)?|"
                         r"^\s*\w{0,14}:?\s*\d{1,2}\s?(am|pm)\b", re.I)
DAY_START = re.compile(r"^\s*(day\s*\w+|night\s*\w+|stop\s*\d|\d{1,2}[.)])", re.I)
BARE_TIME = re.compile(r"^\s*\d{1,2}\s?[:h.]?\s?\d{0,2}\s?(am|pm)?\s*$", re.I)
SCHEDULE = re.compile(r"\b(available|operating from|departures?\s+run|open)\b.*"
                      r"\b(mon|tue|wed|thu|fri|sat|sun|daily)|"
                      r"\b(mon|tue|wed|thu|fri|sat|sun)[a-z]*\s*(?:-|to|–)\s*"
                      r"(mon|tue|wed|thu|fri|sat|sun)", re.I)
ADMIN = re.compile(r"park your vehicle|valet|drop your keys|key drop|"
                   r"your car will be|pick ?up equipment|drop off equipment", re.I)
SELLING = re.compile(r"\bvibes?\b|\bmassive fun\b|\brelaxed,|\bprofessional and\b|"
                     r"well[- ]organi[sz]ed\b|\bsmall groups\b", re.I)
LABEL_LEAD = re.compile(r"^\s*(itinerary|faqs?|min_age|max_age|group_size|duration|"
                        r"meeting_point|highlights|description|pricing)\s*:", re.I)
QUESTION = re.compile(r"^.{5,120}\?\s*$")
INCWORD = (r"includ\w*|inclusi\w*|provided?|provides?|supplied|"
           r"we\s+(will\s+)?(provide|supply)|you\s+get|comes\s+with")
COND = re.compile(r"\bcan supply\b|\bif you (do not|don'?t) have\b|\bcan provide\b|"
                  r"\bis available\b|\bare available\b|\bif required\b|"
                  r"\bcan be (hired|arranged)\b", re.I)


def clean(t):
    return ILLEGAL.sub("", "" if t is None else str(t))[:CELL_LIMIT]


def rows(t):
    return [re.sub(r"^[\s\-–•*]+", "", x).strip()
            for x in str(t or "").split("\n") if x.strip()]


def words(t):
    return set(re.findall(r"[a-z0-9]+", str(t or "").lower()))


def load(path):
    """Parse a batch output file, repairing the model's recurring JSON defect."""
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        cid = r.get("custom_id", "")
        if not cid.endswith("|desc"):
            continue
        pid = cid.split("|")[0]
        try:
            t = r["response"]["body"]["choices"][0]["message"]["content"].strip()
            t = re.sub(r"^```(?:json)?\s*|\s*```$", "", t).strip()
        except Exception:
            out[pid] = None
            continue
        try:
            out[pid] = json.loads(t)
        except json.JSONDecodeError:
            # The model intermittently closes its JSON with a stray `,"` before
            # the brace -- 1-2 products per run, different ones each time. Left
            # unrepaired the product is silently dropped and reads as "empty",
            # which once caused a correct extraction to be reported as data loss.
            try:
                out[pid] = json.loads(re.sub(r',\s*"\s*\}\s*$', "}", t))
                REPAIRED.append(pid)
            except json.JSONDecodeError:
                out[pid] = None
    return out


def inc_heading(raw):
    """A heading announcing an inclusions list -- the word or any synonym."""
    if re.search(r"what_is_included\s*:", raw, re.I):
        return "what_is_included:"
    for l in raw.split("\n"):
        s = l.strip()
        core = s.strip("*#_ ").strip()
        if not core or not re.search(INCWORD, core, re.I):
            continue
        if (s.startswith("#") or s.startswith("**")
                or core.rstrip().endswith((":", ";", "?"))
                or len(core.split()) <= 5):
            return s[:46]
    return None


def judge_itinerary(raw, val):
    r, issues = rows(val), []
    if val.strip():
        if LABEL_LEAD.match(val):
            issues.append("label marker emitted as content")
        if sum(bool(BARE_TIME.match(x)) for x in r) >= 2:
            issues.append("bare-time TIMETABLE, not a route")
        if (sum(bool(CLOCK_START.match(x)) for x in r) < 2
                and sum(bool(DAY_START.match(x)) for x in r) < 2):
            issues.append("no day/time marker on 2+ rows")
        if sum(bool(ADMIN.search(x)) for x in r) >= 2:
            issues.append("booking admin, not the experience")
        if sum(bool(SELLING.search(x)) for x in r) / max(len(r), 1) >= 0.5:
            issues.append("selling points, not a route")
        if any(SCHEDULE.search(x) for x in r):
            issues.append("opening-hours line inside the itinerary")
        if any(re.search(r"\d{1,2}[.)]\s.{6,}\s\d{1,2}[.)]\s", x) for x in r):
            issues.append("two numbered steps merged on one line")
        if {m for m in words(val) - words(raw) if len(m) > 3 and not m.isdigit()}:
            issues.append("words not present in the raw text")
    return issues


def judge_faq(raw, d):
    val = str((d or {}).get(FAQ) or "")
    issues, qs = [], []
    m = re.search(r"faqs?\s*:|##\s*faqs?|frequently asked questions?", raw, re.I)
    if m:
        qs = [re.sub(r"^faqs:\s*", "", x.strip(), flags=re.I)
              for x in raw[m.start():].split("\n") if QUESTION.match(x.strip())]
        if not val.strip():
            issues.append("raw has an FAQ block but the field is empty")
        for q in qs:
            if q[:38].lower() not in val.lower():
                issues.append(f"question missing: {q[:40]}")
            elif [k for k, v in (d or {}).items()
                  if k != FAQ and q[:38].lower() in str(v or "").lower()]:
                issues.append(f"question also in another field: {q[:34]}")
    if val.strip() and {m2 for m2 in words(val) - words(raw)
                        if len(m2) > 3 and not m2.isdigit()}:
        issues.append("words not present in the raw text")
    return issues, len(qs), bool(m)


def judge_included(raw, val):
    issues, h = [], inc_heading(raw)
    if val.strip():
        if not h:
            issues.append("filled but NO inclusions heading in the raw")
        for r in rows(val):
            if COND.search(r):
                issues.append(f"conditional availability kept: {r[:42]}")
                break
        if LABEL_LEAD.match(val):
            issues.append("label marker emitted as content")
        if {m for m in words(val) - words(raw) if len(m) > 3 and not m.isdigit()}:
            issues.append("words not present in the raw text")
    return issues, h


def style(ws, widths=None, wrap=None, fills=None, freeze="A2"):
    ws.freeze_panes = freeze
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = (fills or {}).get(c, HDR)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in (widths or {}).items():
        if c <= ws.max_column:
            ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = (Alignment(wrap_text=True, vertical="top")
                              if wrap and cell.column in wrap
                              else Alignment(vertical="top"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", help="file in output/ (default: newest)")
    args = ap.parse_args()

    # Sort by parsed VERSION, not by name and not by mtime.
    #   by name:  "luna100_v4_8_output" sorts after "luna100_v4_8_3_output"
    #   by mtime: a fresh clone gives every file the same timestamp, so the
    #             newest run is whichever git happened to write last
    def version_key(p):
        m = re.search(r"luna100_v(\d+(?:_\d+)*)_output", p.name)
        return [int(x) for x in m.group(1).split("_")] if m else [0]

    runs = sorted(OUT.glob("luna100_*_output.jsonl"), key=version_key)
    if not runs:
        raise SystemExit("no run output found -- run run_extraction.py first")
    path = OUT / args.run if args.run else runs[-1]
    if not path.exists():
        raise SystemExit(f"{path.name} not found. Available: "
                         + ", ".join(f.name for f in runs))
    print(f"Reviewing {path.name} against the raw supplier text\n")

    out = load(path)
    screen = json.loads((DATA / "luna100_screen_results.json")
                        .read_text(encoding="utf-8"))[MODEL]
    ids = json.loads((DATA / "luna100_products.json")
                     .read_text(encoding="utf-8"))["product_ids"]

    itin, faqs, inc, allr, fails = [], [], [], [], []
    for pid in ids:
        d = out.get(pid)
        raw = screen[pid]["raw_desc"]
        if d is None:
            fails.append(pid)
            continue
        iv = str(d.get(ITIN) or "").strip()
        ii = judge_itinerary(raw, iv)
        fi, nq, hasfaq = judge_faq(raw, d)
        wv = str(d.get(INC) or "").strip()
        wi, wh = judge_included(raw, wv)
        base = {"product_id": pid}
        if iv:
            itin.append({**base, "rows": len(rows(iv)),
                         "VERDICT": "OK" if not ii else "ISSUE",
                         "issues found": "; ".join(ii),
                         "EXTRACTED itinerary": clean(iv),
                         "RAW description (source of truth)": clean(raw),
                         "YOUR VERDICT": "", "YOUR COMMENT": ""})
        if hasfaq or str(d.get(FAQ) or "").strip():
            faqs.append({**base, "questions in raw": nq,
                         "VERDICT": "OK" if not fi else "ISSUE",
                         "issues found": "; ".join(fi),
                         "EXTRACTED faqs": clean(str(d.get(FAQ) or "")),
                         "RAW description (source of truth)": clean(raw),
                         "YOUR VERDICT": "", "YOUR COMMENT": ""})
        inc.append({**base, "filled": "yes" if wv else "no",
                    "heading found in raw": wh or "",
                    "VERDICT": "OK" if not wi else "ISSUE",
                    "issues found": "; ".join(wi),
                    "EXTRACTED What's Included": clean(wv),
                    "About (cross-check)": clean(str(d.get("redo_desc_about") or "")),
                    "RAW description (source of truth)": clean(raw),
                    "YOUR VERDICT": "", "YOUR COMMENT": ""})
        src = words(raw)
        allw = set().union(*(words(v) for v in d.values()))
        allr.append({**base, "itinerary": "yes" if iv else "",
                     "faqs": "yes" if str(d.get(FAQ) or "").strip() else "",
                     "whats included": "yes" if wv else "",
                     "source words kept %": round(100 * len(src & allw)
                                                  / max(len(src), 1), 1),
                     "invented words": ", ".join(sorted(
                         {m for m in allw - src
                          if len(m) > 3 and not m.isdigit()})[:6])})

    def srt(df):
        return df.sort_values(["VERDICT", "product_id"],
                              key=lambda s: s.map({"ISSUE": 0, "OK": 1})
                              if s.name == "VERDICT" else s) if len(df) else df

    itin_df, faq_df = srt(pd.DataFrame(itin)), srt(pd.DataFrame(faqs))
    inc_df, all_df = srt(pd.DataFrame(inc)), pd.DataFrame(allr)

    CHECKS = [
        ("ITINERARY", "Prose accepted via 'then'/'next'", "no day/time marker"),
        ("ITINERARY", "Selling-point list accepted", "selling points"),
        ("ITINERARY", "Booking admin accepted", "booking admin"),
        ("ITINERARY", "Bare-time timetable accepted", "TIMETABLE"),
        ("ITINERARY", "Opening hours inside the field", "opening-hours line"),
        ("ITINERARY", "Raw label emitted as content", "label marker"),
        ("ITINERARY", "Two numbered steps merged", "merged on one line"),
        ("ITINERARY", "Invention", "words not present"),
        ("FAQ", "FAQ block not captured", "field is empty"),
        ("FAQ", "Question dropped", "question missing"),
        ("FAQ", "Question scattered elsewhere", "also in another field"),
        ("WHATS_INCLUDED", "Filled with no supplier heading", "NO inclusions heading"),
        ("WHATS_INCLUDED", "Conditional availability kept", "conditional availability"),
        ("WHATS_INCLUDED", "Invention", "words not present"),
    ]
    crows = []
    for sec, name, needle in CHECKS:
        df = {"ITINERARY": itin_df, "FAQ": faq_df, "WHATS_INCLUDED": inc_df}[sec]
        hit = (df[df["issues found"].str.contains(re.escape(needle), regex=True)]
               if len(df) else df)
        crows.append({"section": sec, "check": name,
                      "status": "PASS" if not len(hit) else "FAIL",
                      "products failing": len(hit),
                      "product ids": " ".join(hit["product_id"]) if len(hit) else ""})
    checks = pd.DataFrame(crows)

    n_pass = int((checks.status == "PASS").sum())
    summary = pd.DataFrame([
        ("Run reviewed", path.name, ""),
        ("Products", len(ids), ""),
        ("Parse failures", len(fails), " ".join(fails) or "none"),
        ("JSON repaired on load", len(REPAIRED), " ".join(REPAIRED) or "none"),
        ("Itinerary filled", len(itin_df), ""),
        ("FAQ filled", int(sum(1 for r in faqs if r["EXTRACTED faqs"].strip())), ""),
        ("What's Included filled", int((inc_df.filled == "yes").sum()), ""),
        ("Checks passing", f"{n_pass} of {len(checks)}", ""),
    ], columns=["measure", "value", "note"])

    read_me = pd.DataFrame([
        ("What this is", "An extraction run checked against the RAW supplier "
         "text. No comparison with any other prompt version."),
        ("Checks sheet", "One row per known issue. PASS means no product shows "
         "it; FAIL lists the product ids."),
        ("Section sheets", "ISSUE rows sort to the top. The raw text sits in the "
         "row so a verdict can be checked without leaving it."),
        ("YOUR VERDICT / YOUR COMMENT", "Blank, for the reviewer."),
        ("Itinerary", "Day-wise or time-wise plan of what happens during the "
         "experience, extracted whole. An activity list with no day and no time "
         "marker is Highlights, not an itinerary."),
        ("FAQs", "Questions with their answers, kept together in source order. A "
         "question is never routed to another field by topic."),
        ("What's Included", "Only filled when the supplier gave a heading "
         "announcing inclusions. Conditional availability -- 'we can supply X if "
         "you do not have your own' -- is not an inclusion."),
    ], columns=["SECTION REVIEW -- run vs raw", "read this first"])

    REPORTS.mkdir(exist_ok=True)
    xlsx = REPORTS / "section_review.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as w:
        read_me.to_excel(w, sheet_name="Read_Me", index=False)
        summary.to_excel(w, sheet_name="Summary", index=False)
        checks.to_excel(w, sheet_name="Checks", index=False)
        itin_df.to_excel(w, sheet_name="Itinerary", index=False)
        faq_df.to_excel(w, sheet_name="FAQ", index=False)
        inc_df.to_excel(w, sheet_name="Whats_Included", index=False)
        all_df.to_excel(w, sheet_name="All_Products", index=False)
        style(w.sheets["Read_Me"], widths={1: 34, 2: 104}, wrap={1, 2})
        style(w.sheets["Summary"], widths={1: 30, 2: 34, 3: 50}, wrap={3})
        ws = w.sheets["Checks"]
        style(ws, widths={1: 16, 2: 38, 3: 9, 4: 10, 5: 40}, wrap={2, 5})
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).fill = (
                GREEN if ws.cell(row=r, column=3).value == "PASS" else RED)
        for sh, vcol, rawcols in (("Itinerary", 3, {5, 6}),
                                  ("FAQ", 3, {5, 6}),
                                  ("Whats_Included", 4, {6, 7, 8})):
            ws = w.sheets[sh]
            n = ws.max_column
            style(ws, widths={1: 11, 2: 10, 3: 9, 4: 9, 5: 60, 6: 60, 7: 60,
                              8: 66, 9: 14, 10: 36},
                  wrap=set(rawcols) | {n},
                  fills={**{c: RAW for c in rawcols},
                         n - 1: REV, n: REV}, freeze="B2")
            ws.auto_filter.ref = ws.dimensions
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=vcol)
                c.fill = GREEN if c.value == "OK" else RED
                ws.cell(row=r, column=n - 1).fill = REV
                ws.cell(row=r, column=n).fill = REV
        style(w.sheets["All_Products"],
              widths={1: 11, 2: 9, 3: 7, 4: 13, 5: 16, 6: 32}, wrap={6})

    md = ["# Section status — run vs raw", "",
          f"**Run:** `{path.name}` · {len(ids)} products · "
          f"{len(fails)} parse failure(s) · {len(REPAIRED)} JSON repaired", "",
          "| Section | Check | Status | Failing | Products |",
          "|---|---|---|---|---|"]
    for _, c in checks.iterrows():
        md.append(f"| {c.section} | {c.check} | "
                  f"{'✅' if c.status == 'PASS' else '❌'} {c.status} | "
                  f"{c['products failing']} | {c['product ids'][:56]} |")
    md += ["", f"**{n_pass} of {len(checks)} checks passing.**", "",
           "## Sections not yet reviewed", "",
           "- [ ] Highlights", "- [ ] What's NOT Included",
           "- [ ] Cancellation policy", "- [ ] Check-in / meeting point",
           "- [ ] Duration / age / group size", "- [ ] Booking-notes side"]
    (REPORTS / "SECTION_STATUS.md").write_text("\n".join(md), encoding="utf-8")

    print(f"  itinerary {len(itin_df)} · faq {len(faq_df)} · "
          f"what's included {int((inc_df.filled == 'yes').sum())}")
    print(f"  checks passing {n_pass} of {len(checks)}")
    for _, c in checks[checks.status == "FAIL"].iterrows():
        print(f"    FAIL  {c.section:<15} {c.check:<34} {c['product ids']}")
    if REPAIRED:
        print(f"  JSON repaired: {' '.join(REPAIRED)}")
    print(f"\nWrote reports/{xlsx.name} and reports/SECTION_STATUS.md")


if __name__ == "__main__":
    main()
