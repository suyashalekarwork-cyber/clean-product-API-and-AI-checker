"""
Build worked_example.xlsx -- one invented product showing every column.

For a manager who wants to see the SHAPE of the output before reading 100 real
rows. The product is made up (a fictional Sydney harbour cruise) so nothing
here can be mistaken for a real supplier's listing, but the structure, the
field names and the heading conventions are exactly those of the real pipeline.

The example is built to exercise the interesting cases rather than the easy
ones:

  - headings the supplier wrote in their own words ("Tour Includes",
    "What You'll Need") that must map to our standard fields
  - a fact stated WITHOUT a heading ("runs for about 3 hours" inside the
    narrative) that is deliberately NOT extracted, because unheaded facts are
    where misfiling happens
  - a "Schedule" block that is departure times, not an itinerary, and so must
    stay in About
  - fields that stay EMPTY because the supplier never wrote them -- the most
    misread part of the output

Two sheets:
  Example_Row      the row exactly as it appears in the real workbook
  Field_By_Field   one row per field, with the source text and the reason,
                   which is the readable version for someone checking the logic

Usage:
    python build_worked_example.py --out <path>
"""
import sys
import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
RAW_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FIELD_FILL = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
GREY = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
AMBER = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

RAW_DESCRIPTION = """duration: 3 Hours
meeting_point: King Street Wharf 3, Darling Harbour
min_age: 8
group_size: Maximum 24 guests
description: Sydney Harbour After Dark

Watch the city light up from the water on our small-group evening cruise. We
cast off as the sun drops behind the Anzac Bridge and follow the shoreline past
Barangaroo, the Opera House and Fort Denison while the skyline turns gold and
then blue. The whole thing runs for about 3 hours at an easy pace, with plenty
of room on the open top deck.

Our skipper has worked this harbour for twenty years and narrates the trip
himself, so expect local stories rather than a recorded commentary.

Tour Includes
- Three-hour harbour cruise with live commentary
- Welcome drink on boarding (sparkling, beer or soft drink)
- Grazing platter of local cheeses and cured meats
- Complimentary tea and coffee all evening

Not Included
- Additional drinks from the licensed bar
- Hotel transfers
- Gratuities

What You'll Need
- A warm layer -- it drops several degrees on the water after sunset
- Flat shoes with grip; the deck can get slippery
- Your booking confirmation, printed or on your phone

Schedule
Departures run Thursday to Sunday at 5:30pm year-round, with an additional
7:00pm sailing during daylight saving.

Cancellation Policy
Full refund if you cancel more than 48 hours before departure. Inside 48 hours
we can move your booking to another sailing subject to availability, but the
fare is non-refundable. Cruises cancelled by us for weather are refunded in
full.

Accessibility
The main deck is step-free and reachable by ramp at most tide levels. The upper
deck is via a staircase only. Please tell us when booking if you use a
wheelchair so we can hold a space near the ramp."""

RAW_BOOKING_NOTES = """Please arrive at King Street Wharf 3 by 5:15pm for a
5:30pm departure. The gangway closes 5 minutes before we sail and late arrivals
cannot be accommodated once we leave the berth.

Bring photo ID if you intend to drink; the bar is licensed and we card anyone
who looks under 25.

Questions? Call the office on the number in your confirmation email between 9am
and 5pm weekdays."""

# field, plain-English name, extracted value, source heading, reason
ROWS = [
    ("redo_desc_about", "About",
     "Sydney Harbour After Dark\n\nWatch the city light up from the water on our "
     "small-group evening cruise. We cast off as the sun drops behind the Anzac "
     "Bridge and follow the shoreline past Barangaroo, the Opera House and Fort "
     "Denison while the skyline turns gold and then blue. The whole thing runs "
     "for about 3 hours at an easy pace, with plenty of room on the open top "
     "deck.\n\nOur skipper has worked this harbour for twenty years and narrates "
     "the trip himself, so expect local stories rather than a recorded "
     "commentary.\n\nSchedule\nDepartures run Thursday to Sunday at 5:30pm "
     "year-round, with an additional 7:00pm sailing during daylight saving.",
     "description:",
     "The narrative, plus anything with no heading of its own. The Schedule "
     "block stays here on purpose: it is WHEN the tour runs, not what happens "
     "during it, so it is not an itinerary."),

    ("redo_desc_highlights", "Highlights", "", "",
     "EMPTY -- the supplier wrote no Highlights heading. We do not invent "
     "highlights by picking attractive sentences out of the description."),

    ("redo_desc_what_included", "What's Included",
     "- Three-hour harbour cruise with live commentary\n"
     "- Welcome drink on boarding (sparkling, beer or soft drink)\n"
     "- Grazing platter of local cheeses and cured meats\n"
     "- Complimentary tea and coffee all evening",
     "Tour Includes",
     "The supplier wrote 'Tour Includes', not 'What's Included'. Both map to "
     "this field -- the heading list accepts the common variants."),

    ("redo_desc_what_excluded", "What's NOT Included",
     "- Additional drinks from the licensed bar\n- Hotel transfers\n- Gratuities",
     "Not Included", "Direct heading match."),

    ("redo_desc_itinerary", "Itinerary", "", "",
     "EMPTY -- there is no Itinerary heading. The Schedule block lists departure "
     "times, which is not a sequence of events during the tour. This is the "
     "field most often filled wrongly, so the bar is deliberately high."),

    ("redo_desc_what_to_bring", "What to Bring",
     "- A warm layer -- it drops several degrees on the water after sunset\n"
     "- Flat shoes with grip; the deck can get slippery\n"
     "- Your booking confirmation, printed or on your phone",
     "What You'll Need",
     "'What You'll Need' is not a standard heading but means the same thing, so "
     "it maps here."),

    ("redo_desc_duration_text", "Duration", "3 Hours", "duration:",
     "Taken from the supplier's own duration field. Note the description ALSO "
     "says 'runs for about 3 hours' -- that phrasing is left in About, because "
     "pulling unheaded facts out is how wrong values get into this field."),

    ("redo_desc_requirements", "Requirements",
     "The main deck is step-free and reachable by ramp at most tide levels. The "
     "upper deck is via a staircase only. Please tell us when booking if you use "
     "a wheelchair so we can hold a space near the ramp.",
     "Accessibility",
     "Accessibility content routes to Requirements under the current mapping."),

    ("redo_desc_cancellation", "Cancellation Policy",
     "Full refund if you cancel more than 48 hours before departure. Inside 48 "
     "hours we can move your booking to another sailing subject to availability, "
     "but the fare is non-refundable. Cruises cancelled by us for weather are "
     "refunded in full.",
     "Cancellation Policy", "Direct heading match."),

    ("redo_desc_check_in", "Check-in", "", "",
     "EMPTY on the description side -- the arrival instructions are in the "
     "booking notes, and the two sides are extracted separately."),

    ("redo_min_age", "Min Age", "8", "min_age:", "Supplier's own field."),
    ("redo_max_age", "Max Age", "", "",
     "EMPTY -- the supplier set no maximum age."),
    ("redo_group_size", "Group Size", "Maximum 24 guests", "group_size:",
     "Supplier's own field."),
    ("redo_meeting_point", "Meeting Point",
     "King Street Wharf 3, Darling Harbour", "meeting_point:",
     "Supplier's own field."),

    ("redo_desc_other", "Other (description)", "", "",
     "EMPTY -- nothing left over. This field only takes pricing, booking "
     "instructions or cross-sell to other tours."),

    ("redo_booking_what_to_bring", "Booking: What to Bring",
     "Bring photo ID if you intend to drink; the bar is licensed and we card "
     "anyone who looks under 25.",
     "(booking notes)",
     "From the booking-notes text. The description side also has a What to Bring "
     "-- the same subject appearing on both sides is correct, because the page "
     "renders them as separate blocks."),

    ("redo_booking_what_not_to_bring", "Booking: What NOT to Bring", "", "",
     "EMPTY."),
    ("redo_booking_inclusions", "Booking: Inclusions", "", "",
     "EMPTY -- inclusions were on the description side."),
    ("redo_booking_location", "Booking: Location", "", "",
     "EMPTY -- the wharf address is in the arrival instruction below rather than "
     "stated separately."),

    ("redo_booking_check_in", "Booking: Check-in",
     "Please arrive at King Street Wharf 3 by 5:15pm for a 5:30pm departure. The "
     "gangway closes 5 minutes before we sail and late arrivals cannot be "
     "accommodated once we leave the berth.",
     "(booking notes)", "Arrival time and place."),

    ("redo_booking_departure_info", "Booking: Departure Info", "", "",
     "EMPTY -- the departure detail sits with the check-in instruction above and "
     "is not repeated here. A sentence belongs in exactly one field."),

    ("redo_booking_itinerary", "Booking: Itinerary", "", "", "EMPTY."),
    ("redo_booking_important_info", "Booking: Important Info", "", "", "EMPTY."),
    ("redo_booking_cancellation", "Booking: Cancellation", "", "",
     "EMPTY -- the cancellation policy was on the description side."),
    ("redo_booking_faqs", "Booking: FAQs", "", "", "EMPTY."),
    ("redo_booking_before_arrival", "Booking: Before Arrival", "", "", "EMPTY."),

    ("redo_booking_contact", "Booking: Contact",
     "Questions? Call the office on the number in your confirmation email "
     "between 9am and 5pm weekdays.",
     "(booking notes)", "Contact instruction."),

    ("redo_booking_other", "Other (booking)", "", "", "EMPTY."),
]


def style(ws, widths, wrap=None, fills=None, freeze="A2"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = (fills or {}).get(c, HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in widths.items():
        if c <= ws.max_column:
            ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = (Alignment(wrap_text=True, vertical="top")
                              if wrap and cell.column in wrap
                              else Alignment(vertical="top"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    n_filled = sum(1 for r in ROWS if r[2].strip())
    src_words = len((RAW_DESCRIPTION + " " + RAW_BOOKING_NOTES).split())
    out_words = sum(len(r[2].split()) for r in ROWS)

    # ---------- Example_Row: exactly the shape of the real workbook ----------
    row = {
        "product_id": "EXAMPLE-001",
        "supplier": "harbourlights-sydney (invented)",
        "words_in_source": src_words,
        "fields_filled": n_filled,
        "text_kept_pct": 100.0,
        "size_ratio": round(out_words / src_words, 3),
        "repeated_sentences": 0,
        "raw_description": RAW_DESCRIPTION,
        "raw_booking_notes": RAW_BOOKING_NOTES,
    }
    for field, nice, value, _, _ in ROWS:
        row[nice] = value
    row["pass/fail"] = ""
    row["comment"] = ""
    ex_df = pd.DataFrame([row])

    # ---------- Field_By_Field: the readable version ----------
    fbf = pd.DataFrame([{
        "field_name_in_data": f,
        "shown_on_page_as": nice,
        "extracted_value": val,
        "came_from_heading": src if src else "(no heading)",
        "why": why,
        "status": "FILLED" if val.strip() else "EMPTY (correct)",
    } for f, nice, val, src, why in ROWS])

    # ---------- Read_Me ----------
    readme = pd.DataFrame([
        ("WHAT THIS IS", ""),
        ("", "An invented product, used to show the SHAPE of the output. No "
             "real supplier is involved. The field names, headings and rules "
             "are the real ones."),
        ("", ""),
        ("THE TWO SHEETS", ""),
        ("Example_Row", "The row exactly as it appears in the real 100-product "
                        "workbook: supplier's raw text in the yellow columns, "
                        "then every extracted field to the right."),
        ("Field_By_Field", "The same information one field per row, with the "
                           "heading it came from and why. Easier to read if you "
                           "are checking the logic rather than the layout."),
        ("", ""),
        ("THE FOUR THINGS WORTH NOTICING", ""),
        ("1. Supplier headings vary",
         "This supplier wrote 'Tour Includes' and 'What You'll Need'. We accept "
         "the common variants and map them to the standard fields."),
        ("2. Empty is a real answer",
         f"{len(ROWS) - n_filled} of {len(ROWS)} fields are empty here, because "
         f"the supplier never wrote them. That is correct output, not a "
         f"failure. Most tours have no itinerary and no FAQ."),
        ("3. Unheaded facts are left alone",
         "The description says the cruise 'runs for about 3 hours'. We do NOT "
         "use that to fill Duration -- we use the supplier's own duration "
         "field. Pulling facts out of prose is how wrong values get in: 'be "
         "ready 10 minutes before departure' would otherwise set Duration to 10 "
         "minutes."),
        ("4. Schedule is not an Itinerary",
         "'Departures run Thursday to Sunday at 5:30pm' is WHEN the tour runs, "
         "not what happens during it. It stays in About. Itinerary is the field "
         "most often filled wrongly, so it requires a real sequence."),
        ("", ""),
        ("HOW TO CHECK ANY PRODUCT", ""),
        ("", "Read the yellow raw-text column, then read left to right. Every "
             "word in an extracted field should be findable in the raw text. "
             "Nothing should appear in two fields."),
    ], columns=["", " "])

    out_path = Path(args.out)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        readme.to_excel(w, sheet_name="Read_Me", index=False)
        ex_df.to_excel(w, sheet_name="Example_Row", index=False)
        fbf.to_excel(w, sheet_name="Field_By_Field", index=False)

        ws = w.sheets["Read_Me"]
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 105
        for r in ws.iter_rows(min_row=1):
            for cell in r:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r[0].value and not str(r[1].value or "").strip():
                r[0].font = Font(bold=True, size=12)
                r[0].fill = HEADER_FILL

        cols = list(ex_df.columns)
        i_raw1 = cols.index("raw_description") + 1
        i_raw2 = cols.index("raw_booking_notes") + 1
        i_pf = cols.index("pass/fail") + 1
        fills = {i_raw1: RAW_FILL, i_raw2: RAW_FILL}
        for c in range(i_raw2 + 1, i_pf):
            fills[c] = FIELD_FILL
        widths = {1: 14, 2: 26, 3: 10, 4: 9, 5: 11, 6: 10, 7: 12,
                  i_raw1: 80, i_raw2: 60}
        for c in range(i_raw2 + 1, len(cols) + 1):
            widths[c] = 46
        style(w.sheets["Example_Row"], widths,
              wrap=set(range(i_raw1, len(cols) + 1)), fills=fills)
        w.sheets["Example_Row"].row_dimensions[2].height = 400

        ws = w.sheets["Field_By_Field"]
        style(ws, {1: 30, 2: 26, 3: 60, 4: 22, 5: 70, 6: 18}, wrap={3, 5})
        for r in range(2, ws.max_row + 1):
            filled = str(ws.cell(r, 6).value or "").startswith("FILLED")
            ws.cell(r, 6).fill = GREEN if filled else GREY
            if not filled:
                ws.cell(r, 5).fill = AMBER

    print("=" * 74)
    print("WORKED EXAMPLE")
    print("=" * 74)
    print(f"  invented product : EXAMPLE-001 (Sydney Harbour After Dark)")
    print(f"  source words     : {src_words}")
    print(f"  fields filled    : {n_filled} of {len(ROWS)}")
    print(f"  fields empty     : {len(ROWS) - n_filled} (all correct -- supplier "
          f"never wrote them)")
    print(f"  size ratio       : {out_words / src_words:.3f}")
    print(f"\nWrote {out_path.name}")


if __name__ == "__main__":
    main()
